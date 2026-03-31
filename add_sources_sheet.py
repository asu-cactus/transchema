"""
Adds a 'Sources' sheet to both result spreadsheets with:
  - Log directory path
  - CSV file paths (average_multi_step, critique, final_critique)
  - Python command used to run each experiment
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_LOG = "logs-auto-suggest-llm-21-04"

# ── Styles ──────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="BDD7EE")
PEND_FILL = PatternFill("solid", fgColor="FFEB9C")
PEND_FONT = Font(color="9C5700")
BOLD      = Font(bold=True)
WRAP      = Alignment(wrap_text=True, vertical="top")
TOP       = Alignment(vertical="top")

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def set_header(ws, row, cols):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = BOLD
        cell.fill = HDR_FILL
        cell.alignment = TOP
        cell.border = thin_border()

def set_cell(ws, row, col, value, pending=False, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = WRAP if wrap else TOP
    cell.border = thin_border()
    if pending:
        cell.fill = PEND_FILL
        cell.font = PEND_FONT
    return cell

def autofit(ws, min_w=12, max_w=80):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                best = max(best, min(max(len(l) for l in lines), max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = best + 2


# ── Spreadsheet 1: results_gpt41_mini_o4_mini.xlsx ──────────────────────────

RUNS_41 = [
    # (label, dir_suffix, length, judge, start, end)
    ("gpt-4.1-mini", "gt",               "L1", "gt",               "judge_exp_gt_gpt41_mini_L1_20260325_135133",              0,  100),
    ("gpt-4.1-mini", "llm_score_hybrid", "L1", "llm_score_hybrid", "judge_exp_llm_score_hybrid_gpt41_mini_L1_20260325_145701", 0,  100),
    ("gpt-4.1-mini", "gt",               "L4", "gt",               "judge_exp_gt_gpt41_mini_L4_20260325_160016",              15, 100),
    ("gpt-4.1-mini", "llm_score_hybrid", "L4", "llm_score_hybrid", "judge_exp_llm_score_hybrid_gpt41_mini_L4_20260325_184305", 15, 100),
    ("gpt-4.1-mini", "gt",               "L9", "gt",               "judge_exp_gt_gpt41_mini_L9_20260325_215200",              0,  100),
    ("gpt-4.1-mini", "llm_score_hybrid", "L9", "llm_score_hybrid", "judge_exp_llm_score_hybrid_gpt41_mini_L9_20260326_012146", 0,  100),
    ("o4-mini",      "gt",               "L9", "gt",               "judge_exp_gt_o4_mini_L9_20260325_135041",                0,  100),
    ("o4-mini",      "llm_score_hybrid", "L9", "llm_score_hybrid", "judge_exp_llm_score_hybrid_o4_mini_L9_20260325_175242",   0,  100),
    ("o4-mini",      "gt",               "L4", "gt",               "judge_exp_gt_o4_mini_L4_20260325_215305",                15, 100),
    # pending
    ("o4-mini",      "llm_score_hybrid", "L4", "llm_score_hybrid", None, 15, 100),
    ("o4-mini",      "gt",               "L1", "gt",               None,  0, 100),
    ("o4-mini",      "llm_score_hybrid", "L1", "llm_score_hybrid", None,  0, 100),
]

def make_cmd(model, judge, len_id, start, end):
    len_num = len_id.replace("L", "")
    return (
        f"python3 critique_data.py \\\n"
        f"  --len_id {len_num} \\\n"
        f"  --target_id {start} \\\n"
        f"  --max_target_id {end} \\\n"
        f"  --model {model} \\\n"
        f"  --judge {judge} \\\n"
        f"  --validation autopipeline \\\n"
        f"  --benchmark github \\\n"
        f"  --source-length 3 \\\n"
        f"  --target-length 3 \\\n"
        f"  --experiment-name judge_exp_{judge}_{model.replace('-','_')}_{len_id}"
    )

wb = openpyxl.load_workbook("results_gpt41_mini_o4_mini.xlsx")

if "Sources" in wb.sheetnames:
    del wb["Sources"]
ws = wb.create_sheet("Sources")

COLS = ["Model", "Judge", "Length", "Status",
        "Experiment Name", "Log Directory",
        "average_multi_step.csv", "critique.csv", "final_critique.csv",
        "Python Command"]

# Title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
title = ws.cell(row=1, column=1, value="Source Paths & Commands — gpt-4.1-mini & o4-mini Experiments")
title.font = Font(bold=True, size=13)
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 22

set_header(ws, 2, COLS)

for r, (model, _, length, judge, exp_dir, start, end) in enumerate(RUNS_41, start=3):
    pending = exp_dir is None
    status  = "PENDING" if pending else "COMPLETE"

    if pending:
        log_dir = f"{BASE_LOG}/<not yet created>"
        ms_csv  = "(pending)"
        cr_csv  = "(pending)"
        fc_csv  = "(pending)"
        exp_name = f"judge_exp_{judge}_{model.replace('-','_')}_{length}"
    else:
        log_dir  = f"{BASE_LOG}/{exp_dir}"
        ms_csv   = f"{log_dir}/results/average_multi_step.csv"
        cr_csv   = f"{log_dir}/results/critique.csv"
        fc_csv   = f"{log_dir}/results/final_critique.csv"
        exp_name = exp_dir.rsplit("_202", 1)[0]   # strip timestamp

    cmd = make_cmd(model, judge, length, start, end)

    set_cell(ws, r, 1,  model,    pending)
    set_cell(ws, r, 2,  judge,    pending)
    set_cell(ws, r, 3,  length,   pending)
    set_cell(ws, r, 4,  status,   pending)
    set_cell(ws, r, 5,  exp_name, pending)
    set_cell(ws, r, 6,  log_dir,  pending)
    set_cell(ws, r, 7,  ms_csv,   pending)
    set_cell(ws, r, 8,  cr_csv,   pending)
    set_cell(ws, r, 9,  fc_csv,   pending)
    set_cell(ws, r, 10, cmd,      pending, wrap=True)
    ws.row_dimensions[r].height = 90

ws.freeze_panes = "A3"
ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 9
ws.column_dimensions["D"].width = 11
ws.column_dimensions["E"].width = 42
ws.column_dimensions["F"].width = 60
ws.column_dimensions["G"].width = 65
ws.column_dimensions["H"].width = 65
ws.column_dimensions["I"].width = 65
ws.column_dimensions["J"].width = 70

wb.save("results_gpt41_mini_o4_mini.xlsx")
print("Saved results_gpt41_mini_o4_mini.xlsx")


# ── Spreadsheet 2: results_memory_granularity.xlsx ──────────────────────────

RUNS_MEM = [
    # (mode, dir_suffix, cases_desc)
    ("summary",  "memory_granularity_summary_20260325_141429_v2_20260325_142209",
     "L1: 0-9 (pilot)", False),
    ("response", "memory_granularity_response_20260325_141429_v2_20260325_142446",
     "L1: 0-9 (pilot)", False),
    ("qa",       "memory_granularity_qa_20260325_141429_v2_20260325_142833",
     "L1: 0-9 (pilot, token limit hit for 8/10 cases)", False),
    ("response", None, "L1: 0-99, L4: 18-99",  True),
    ("qa",       None, "L1: 0-99, L4: 18-99",  True),
]

def make_mem_cmd(mode, cases_desc):
    if "0-99" in cases_desc:
        cases_note = "(see run_mem_granularity.sh for full case list)"
        return (
            f"# See run_mem_granularity.sh\n"
            f"python3 critique_data.py \\\n"
            f"  --model gpt-4.1-mini \\\n"
            f"  --judge gt \\\n"
            f"  --cases $L1_CASES $L4_CASES \\\n"
            f"  --memory_granularity {mode} \\\n"
            f"  --experiment-name mem_gran_{mode}_$TS"
        )
    else:
        return (
            f"python3 critique_data.py \\\n"
            f"  --model gpt-4.1-mini \\\n"
            f"  --judge gt \\\n"
            f"  --cases 1_0 1_1 1_2 1_3 1_4 1_5 1_6 1_7 1_8 1_9 \\\n"
            f"  --memory_granularity {mode} \\\n"
            f"  --experiment-name mem_gran_{mode}_<timestamp>"
        )

wb2 = openpyxl.load_workbook("results_memory_granularity.xlsx")

if "Sources" in wb2.sheetnames:
    del wb2["Sources"]
ws2 = wb2.create_sheet("Sources")

COLS2 = ["Mode", "Cases", "Status",
         "Experiment Name", "Log Directory",
         "average_multi_step.csv", "critique.csv", "final_critique.csv",
         "Python Command"]

ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS2))
title2 = ws2.cell(row=1, column=1, value="Source Paths & Commands — Memory Granularity Experiments")
title2.font = Font(bold=True, size=13)
title2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

set_header(ws2, 2, COLS2)

for r, (mode, exp_dir, cases_desc, pending) in enumerate(RUNS_MEM, start=3):
    status = "PENDING" if pending else "COMPLETE"

    if pending:
        log_dir  = f"{BASE_LOG}/<not yet created>"
        ms_csv   = "(pending)"
        cr_csv   = "(pending)"
        fc_csv   = "(pending)"
        exp_name = f"mem_gran_{mode}_<timestamp>"
    else:
        log_dir  = f"{BASE_LOG}/{exp_dir}"
        ms_csv   = f"{log_dir}/results/average_multi_step.csv"
        cr_csv   = f"{log_dir}/results/critique.csv"
        fc_csv   = f"{log_dir}/results/final_critique.csv"
        exp_name = exp_dir.rsplit("_202", 1)[0]

    cmd = make_mem_cmd(mode, cases_desc)

    set_cell(ws2, r, 1, mode,       pending)
    set_cell(ws2, r, 2, cases_desc, pending)
    set_cell(ws2, r, 3, status,     pending)
    set_cell(ws2, r, 4, exp_name,   pending)
    set_cell(ws2, r, 5, log_dir,    pending)
    set_cell(ws2, r, 6, ms_csv,     pending)
    set_cell(ws2, r, 7, cr_csv,     pending)
    set_cell(ws2, r, 8, fc_csv,     pending)
    set_cell(ws2, r, 9, cmd,        pending, wrap=True)
    ws2.row_dimensions[r].height = 95

ws2.freeze_panes = "A3"
ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 11
ws2.column_dimensions["D"].width = 38
ws2.column_dimensions["E"].width = 60
ws2.column_dimensions["F"].width = 65
ws2.column_dimensions["G"].width = 65
ws2.column_dimensions["H"].width = 65
ws2.column_dimensions["I"].width = 70

wb2.save("results_memory_granularity.xlsx")
print("Saved results_memory_granularity.xlsx")
