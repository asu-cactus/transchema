"""
update_mcts_results.py

Patches the MCTS is_correct values in results_analysis.xlsx using
the output from revalidate_mcts.py (revalidate_mcts_results.csv).

Only the MCTS column in the L4 / L9 per-case sheets and the
corresponding rows in the Overall sheet are touched.

Run from ~/transchema/:
    python update_mcts_results.py
    python update_mcts_results.py --revalidate-csv my_results.csv
    python update_mcts_results.py --revalidate-csv my_results.csv --xlsx results_analysis.xlsx
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ── Defaults ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
DEFAULT_REVALIDATE_CSV = ROOT / "revalidate_mcts_results.csv"
DEFAULT_XLSX           = ROOT / "results_analysis.xlsx"

# Column header used by compile_results.py for MCTS
MCTS_COL_LABEL = "Operator-driven\nMCTS"

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--revalidate-csv", default=str(DEFAULT_REVALIDATE_CSV),
                    help="Path to revalidate_mcts_results.csv")
parser.add_argument("--xlsx", default=str(DEFAULT_XLSX),
                    help="Path to results_analysis.xlsx (edited in place)")
args = parser.parse_args()

revalidate_csv = Path(args.revalidate_csv)
xlsx_path      = Path(args.xlsx)

# ── Load revalidation results ─────────────────────────────────────────────────
df_rev = pd.read_csv(revalidate_csv)

# case_id in revalidate is "length4_32"; xlsx uses "4_32"
def strip_length_prefix(cid: str) -> str:
    if str(cid).startswith("length"):
        return str(cid)[len("length"):]
    return str(cid)

df_rev["case_id_short"] = df_rev["case_id"].apply(strip_length_prefix)
df_rev["is_correct"] = df_rev["is_correct"].map(
    lambda v: True if str(v).strip().lower() in ("true", "1") else False
)

# Map: short case_id → is_correct
corrections: dict[str, bool] = dict(zip(df_rev["case_id_short"], df_rev["is_correct"]))
lengths_in_data = set(df_rev["case_id_short"].str.split("_").str[0])

print(f"Loaded {len(corrections)} cases from {revalidate_csv.name}")
print(f"Lengths present: {sorted(lengths_in_data)}")
print(f"Correct after revalidation: {sum(corrections.values())} / {len(corrections)}")

# ── Open workbook ─────────────────────────────────────────────────────────────
wb = load_workbook(xlsx_path)

# ── Helper: find column index by header value ─────────────────────────────────
def find_col(ws, header: str) -> int | None:
    for cell in ws[1]:
        if cell.value and str(cell.value).strip() == header.strip():
            return cell.column
    return None

# ── Patch per-length sheets ───────────────────────────────────────────────────
patched_total = 0

for sheet_name, target_len_str in [("L4", "4"), ("L9", "9")]:
    if target_len_str not in lengths_in_data:
        print(f"\nSkipping {sheet_name} — no cases for length {target_len_str} in revalidate CSV")
        continue

    if sheet_name not in wb.sheetnames:
        print(f"\nSkipping {sheet_name} — sheet not found in xlsx")
        continue

    ws = wb[sheet_name]
    mcts_col = find_col(ws, MCTS_COL_LABEL)
    case_col  = find_col(ws, "Case ID")

    if mcts_col is None:
        print(f"\nWARNING: '{MCTS_COL_LABEL}' column not found in {sheet_name} sheet")
        continue
    if case_col is None:
        print(f"\nWARNING: 'Case ID' column not found in {sheet_name} sheet")
        continue

    patched = 0
    for row in ws.iter_rows(min_row=2):
        case_id = str(row[case_col - 1].value) if row[case_col - 1].value is not None else ""
        if case_id in corrections:
            old_val = row[mcts_col - 1].value
            new_val = corrections[case_id]
            row[mcts_col - 1].value = new_val
            if str(old_val).strip().lower() != str(new_val).lower():
                print(f"  {sheet_name} {case_id}: {old_val} → {new_val}")
            patched += 1

    print(f"\n{sheet_name}: patched {patched} / {ws.max_row - 1} rows")
    patched_total += patched

# ── Patch Overall sheet ───────────────────────────────────────────────────────
# Recompute Correct count for MCTS rows from the now-patched per-length sheets

if "Overall" in wb.sheetnames:
    ws_overall = wb["Overall"]

    # Read header row
    headers = [c.value for c in ws_overall[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    for row in ws_overall.iter_rows(min_row=2):
        decision_cell = row[col_idx["Decision"] - 1]
        planning_cell = row[col_idx["Planning"] - 1]
        length_cell   = row[col_idx["Length"]   - 1]

        if (str(decision_cell.value) == "Operator-driven"
                and str(planning_cell.value) == "MCTS"):

            length_str = str(length_cell.value)  # e.g. "L4"
            len_num    = length_str.lstrip("L")   # "4"

            if len_num not in lengths_in_data:
                continue

            # Count corrects from the corrections dict for this length
            correct_count = sum(
                v for k, v in corrections.items()
                if k.split("_")[0] == len_num
            )
            total_count = sum(1 for k in corrections if k.split("_")[0] == len_num)

            old_correct = row[col_idx["Correct"] - 1].value
            row[col_idx["Correct"] - 1].value = correct_count
            row[col_idx["Cases Evaluated"] - 1].value = total_count

            print(f"\nOverall MCTS {length_str}: Correct {old_correct} → {correct_count} / {total_count}")

# ── Auto-fit columns ──────────────────────────────────────────────────────────
for sheet in wb.worksheets:
    for col_cells in sheet.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        sheet.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 60)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(xlsx_path)
print(f"\nDone. Patched {patched_total} cells. Saved to: {xlsx_path}")
