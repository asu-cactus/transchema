"""
Script to add "AgentFlow Prompt Designs - Operator" sheet to Prompt Design.xlsx.
Run from transchema/ root:  python3 add_agentflow_sheet.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

XLSX_PATH = "Prompt Design.xlsx"
SHEET_NAME = "AF Prompt Designs - Operator"

# ---------------------------------------------------------------------------
# Row data: (AgentFlow Component, Section Type, Current AgentFlow Text,
#            Maps To (Multistep), Match Status)
# ---------------------------------------------------------------------------
ROWS = [
    # ── Planner ANALYZE_QUERY ──────────────────────────────────────────────
    (
        "Planner: ANALYZE_QUERY (operator mode)",
        "Task Description",
        "Analyze the given query to determine necessary skills and tools. "
        "You can ONLY suggest operations and tools that are in the available tools list.",
        "No multistep equivalent (multistep uses hardcoded allowed_operation_list)",
        "No Multistep Equivalent",
    ),
    (
        "Planner: ANALYZE_QUERY (operator mode)",
        "Context fields",
        "Inputs: Query, Available tools, Metadata for tools",
        "No multistep equivalent",
        "No Multistep Equivalent",
    ),
    (
        "Planner: ANALYZE_QUERY (operator mode)",
        "Output format",
        "Summary of query, lists of skills and tools with explanations, additional considerations.",
        "No multistep equivalent",
        "No Multistep Equivalent",
    ),
    # ── Planner GENERATE_NEXT_STEP ─────────────────────────────────────────
    (
        "Planner: GENERATE_NEXT_STEP (operator mode)",
        "Task Description",
        "Task: Determine the optimal next step to build the data transformation pipeline operator by operator. "
        "You are operating at OPERATOR granularity. "
        "CRITICAL CONSTRAINT: You can ONLY use tools from the available tools list.",
        "get_next_operator_prompt — Task: 'what operation should be performed next?' "
        "Allowed Operations listed; Operation History; source/target schema + examples.",
        "Needs Update → Done",
    ),
    (
        "Planner: GENERATE_NEXT_STEP (operator mode)",
        "Tool Descriptions",
        "Configure_Join_Operator_Tool, Configure_Union_Operator_Tool, "
        "Configure_GroupBy_Aggregate_Operator_Tool, Add_Pivot_Tool, Add_Unpivot_Tool, "
        "Code_Gen_And_Score_Tool, Critique_Pipeline_Tool — each with usage guidance.",
        "Allowed Operations list in multistep (JOIN, UNION, GROUP_BY, AGGREGATE, etc.)",
        "Needs Update → Done",
    ),
    (
        "Planner: GENERATE_NEXT_STEP (operator mode)",
        "Domain Hints (NEW — added in this update)",
        "=== DOMAIN HINTS FOR OPERATOR SELECTION ===\n"
        "[get_hints_section(NEXT_OPERATOR_HINT_IDS, fmt='bullet')]\n"
        "Hint IDs: [1,2,3,4,5,6,9,11,16]",
        "get_next_operator_prompt — static_hints block using NEXT_OPERATOR_HINT_IDS",
        "Needs Update → Done",
    ),
    (
        "Planner: GENERATE_NEXT_STEP (operator mode)",
        "Context fields",
        "Query, Query Analysis, Available Tools, Toolbox Metadata, "
        "Previous Steps and Their Results, Current Step, Remaining Steps.",
        "Source Information, Operation History, Target Schema, Target Examples",
        "Needs Update → Done",
    ),
    (
        "Planner: GENERATE_NEXT_STEP (operator mode)",
        "Response format",
        "Justification → Context → Sub-Goal → Tool Name (in that order, nothing after).",
        "Final answer in $OPERATOR$ quotes",
        "Needs Update → Done",
    ),
    # ── Configure_Join_Operator_Tool ───────────────────────────────────────
    (
        "Configure_Join_Operator_Tool",
        "Task Description",
        "You are generating a data-pipeline to transform multiple source tables to target table "
        "and you need to answer 'what tables should be joined and at which columns?'",
        "get_join_prompt — identical task description",
        "Identical",
    ),
    (
        "Configure_Join_Operator_Tool",
        "Case Info format",
        "transformation_case_info (contains target table name, schema, examples, "
        "multi source information, operation history)",
        "Separate args: target_data_name, target_data_schema, target_samples, "
        "source_information, operation_history, hints, fd_hints",
        "Needs Update → Done",
    ),
    (
        "Configure_Join_Operator_Tool",
        "Task+Case Specific Hints",
        "[get_hints_section(JOIN_HINT_IDS, fmt='bullet')]\nHint IDs: [7,8,9]",
        "get_join_prompt — get_hints_section(JOIN_HINT_IDS, fmt='bullet')",
        "Needs Update → Done",
    ),
    (
        "Configure_Join_Operator_Tool",
        "Return format",
        "[ [table1, table2], [table1.col1, table2.col1], ... ] within $ quotes",
        "Same format",
        "Identical",
    ),
    # ── Configure_Union_Operator_Tool ──────────────────────────────────────
    (
        "Configure_Union_Operator_Tool",
        "Task Description",
        "You are generating a data-pipeline to transform multiple source tables to target table "
        "and you need to answer 'what tables should be Union-ed?'. Take this decision based on "
        "'Operation History', few shot examples, 'Source' and 'Target' (table schema as well as "
        "the examples) information provided below.",
        "get_union_prompt — identical task description (now aligned)",
        "Needs Update → Done",
    ),
    (
        "Configure_Union_Operator_Tool",
        "Case Info format",
        "transformation_case_info",
        "Separate args: target_data_name, target_data_schema, target_samples, "
        "source_information, operation_history",
        "Needs Update → Done",
    ),
    (
        "Configure_Union_Operator_Tool",
        "Task Specific General Hints",
        "No domain hints (none in mapping for Union)",
        "No domain hints in get_union_prompt either",
        "Identical",
    ),
    (
        "Configure_Union_Operator_Tool",
        "Return format",
        "[$table1$, $table2$, ...]",
        "Same format",
        "Identical",
    ),
    # ── Configure_GroupBy_Aggregate_Operator_Tool ──────────────────────────
    (
        "Configure_GroupBy_Aggregate_Operator_Tool",
        "Task Description",
        "You are generating a data-pipeline to transform multiple source tables to target table "
        "and you need to answer: 1. Which columns for Group By? 2. Which columns Aggregated? "
        "3. Which Aggregation functions?",
        "get_group_by_aggregate_prompt — identical task description",
        "Identical",
    ),
    (
        "Configure_GroupBy_Aggregate_Operator_Tool",
        "Case Info format",
        "transformation_case_info",
        "Separate args: target_data_name, target_data_schema, target_samples, "
        "source_information, operation_history, hints, fd_hints",
        "Needs Update → Done",
    ),
    (
        "Configure_GroupBy_Aggregate_Operator_Tool",
        "Task+Case Specific Hints",
        "[get_hints_section(GROUPBY_AGG_HINT_IDS, fmt='bullet')]\nHint IDs: [10,14,16,17,18,20,21]",
        "get_group_by_aggregate_prompt — get_hints_section(GROUPBY_AGG_HINT_IDS, fmt='bullet')",
        "Needs Update → Done",
    ),
    (
        "Configure_GroupBy_Aggregate_Operator_Tool",
        "Return format",
        '"group_by" = [...], "aggregations" = [agg_func(col), ...]',
        "Same format",
        "Identical",
    ),
    # ── Code_Gen_And_Score_Tool ────────────────────────────────────────────
    (
        "Code_Gen_And_Score_Tool",
        "Task Description",
        "You are generating executable Python code at runtime. Please generate a Python script "
        "to convert multiple source tables to the format of the target table and STRICTLY follow "
        "the sequence of the operations mentioned in the action history below.",
        "get_python_script_simple — identical task description",
        "Identical",
    ),
    (
        "Code_Gen_And_Score_Tool",
        "Context format",
        "Transformation Context: {query}\nAction History (follow this sequence): {memory_actions}",
        "Operation History: {operation_history}\nSource Information: {source_information_with_location}",
        "Needs Update → Done",
    ),
    (
        "Code_Gen_And_Score_Tool",
        "Task+Case Specific Hints",
        "[get_hints_section(PYTHON_SCRIPT_HINT_IDS, fmt='bullet')]\n"
        "Hint IDs: [1,2,3,4,5,10,11,16,17,24,27,28,29,30,31,32]",
        "get_python_script_simple — get_hints_section(PYTHON_SCRIPT_HINT_IDS, fmt='bullet')",
        "Needs Update → Done",
    ),
    (
        "Code_Gen_And_Score_Tool",
        "Return format",
        "```Python ... ``` + error string: Errors in previous attempts: {error_string}",
        "Same format",
        "Identical",
    ),
    # ── Critique_Pipeline_Tool ─────────────────────────────────────────────
    (
        "Critique_Pipeline_Tool",
        "Task Description",
        "You are critiquing a data-pipeline that transforms multiple source tables to follow the "
        "schema of the target table. Review the current pipeline and suggest corrections.",
        "critique.py query_generator — similar task description",
        "Needs Update → Done",
    ),
    (
        "Critique_Pipeline_Tool",
        "Case Info format",
        "transformation_case_info (target table, sources, operation history)",
        "critique.py — query contains target schema, examples, source info, operation history",
        "Needs Update → Done",
    ),
    (
        "Critique_Pipeline_Tool",
        "Per-operator config instructions",
        "--For Join: tables + join attributes.\n"
        "--For Union: tables to union.\n"
        "--For GroupBy: Group By attributes (leftmost non-float unique cols).\n"
        "--For Aggregate: aggregation function + column (not in GroupBy).",
        "critique.py — similar per-operator guidance in query_generator",
        "Needs Update → Done",
    ),
    (
        "Critique_Pipeline_Tool",
        "Task+Case Specific Hints",
        "[get_hints_section(CRITIQUE_HINT_IDS, fmt='numbered')]\n"
        "Hint IDs: [4,5,7,8,9,10,11,12,13,15,16,18,19,22,23,24,25,26,28]",
        "critique.py — get_hints_section(CRITIQUE_HINT_IDS, fmt='numbered')",
        "Needs Update → Done",
    ),
    (
        "Critique_Pipeline_Tool",
        "Return format",
        "$CRITIQUE: <detailed critique and corrected pipeline>$",
        "critique.py — similar $CRITIQUE: ...$ delimited format",
        "Needs Update → Done",
    ),
    # ── Verifier.verificate_context ────────────────────────────────────────
    (
        "Verifier: verificate_context (operator mode)",
        "Task Description",
        "Task: Evaluate whether the operator-driven transformation pipeline has produced a correct "
        "and complete result, or whether more steps are needed.",
        "No multistep equivalent",
        "No Multistep Equivalent",
    ),
    (
        "Verifier: verificate_context (operator mode)",
        "Context fields",
        "Query, Available Tools, Toolbox Metadata, Initial Analysis, "
        "Memory (Tools Used & Results)",
        "No multistep equivalent",
        "No Multistep Equivalent",
    ),
    (
        "Verifier: verificate_context (operator mode)",
        "Stop/Continue logic",
        "1. Review operator configs in memory.\n"
        "2. Identify whether all operators configured to produce target schema.\n"
        "3. Code and Score Check: if Code_Gen_And_Score_Tool result with score=1.0 and "
        "no missed items → STOP; else CONTINUE.\n"
        "Response: 2-3 sentences ending with 'Conclusion: STOP' or 'Conclusion: CONTINUE'.",
        "No multistep equivalent",
        "No Multistep Equivalent",
    ),
]

# ---------------------------------------------------------------------------
# Match status fill colors
# ---------------------------------------------------------------------------
FILL_COLORS = {
    "Identical": "C6EFCE",              # green
    "Needs Update → Done": "FFEB9C",    # yellow
    "No Multistep Equivalent": "BDD7EE",  # blue
    "No AgentFlow Equivalent": "FCE4D6",  # orange
}

HEADER_FILL = "4472C4"  # dark blue
HEADER_FONT_COLOR = "FFFFFF"


def build_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet(title=SHEET_NAME)

    headers = [
        "AgentFlow Component",
        "Section Type",
        "Current AgentFlow Text",
        "Maps To (Multistep)",
        "Match Status",
    ]

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Write data rows
    for row_idx, row_data in enumerate(ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Color the Match Status cell (column 5)
        status = row_data[4]
        if status in FILL_COLORS:
            ws.cell(row=row_idx, column=5).fill = PatternFill(
                "solid", fgColor=FILL_COLORS[status]
            )

    # Set column widths
    col_widths = [36, 32, 70, 60, 26]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"


LG_SHEET_NAME = "LG Prompt Designs - MCTS"

LG_ROWS = [
    # ── mcts_expand: Task Description ─────────────────────────────────────
    (
        "mcts_expand",
        "Task Description",
        "You are building a data-pipeline to transform multiple source tables into the target table. "
        "Given the current 'Operation History', propose up to k ALTERNATIVE candidates for the SINGLE "
        "NEXT operation step, ranked from most to least promising. Each candidate is an INDEPENDENT "
        "ALTERNATIVE — NOT sequential. For EVERY candidate specify BOTH operator type AND its full "
        "configuration.",
        "get_next_operator_prompt + get_join_prompt + get_group_by_aggregate_prompt (combined into one call)",
        "Identical",
    ),
    # ── mcts_expand: Context fields ────────────────────────────────────────
    (
        "mcts_expand",
        "Context fields",
        "Allowed Operations, Operation History, Target Table Name, Target Schema, Target Examples, "
        "Source Information, FD Hints (fd_hints), k (number of candidates)",
        "get_next_operator_prompt: target_data_name, target_data_schema, target_samples, "
        "source_information, operation_history, allowed_operation_list, hints, fd_hints",
        "Identical",
    ),
    # ── mcts_expand: Config Rules — JOIN ──────────────────────────────────
    (
        "mcts_expand",
        "Config Rules — JOIN",
        "[get_hints_section(JOIN_HINT_IDS, fmt='bullet')]\nHint IDs: [7,8,9]\n"
        "+ non-hint line: 'You should only use columns that actually exist in the source tables.'\n"
        "+ Format: TABLES: [[t1, t2]], COLUMNS: [[t1.col_a, t2.col_b], ...]",
        "get_join_prompt — get_hints_section(JOIN_HINT_IDS, fmt='bullet')",
        "Needs Update → Done",
    ),
    # ── mcts_expand: Config Rules — UNION ─────────────────────────────────
    (
        "mcts_expand",
        "Config Rules — UNION",
        "Only union tables with EXACTLY the same column names; do NOT rename. "
        "If tables have different schemas, use JOIN instead of UNION.\n"
        "Format: TABLES: [table1, table2, table3, ...]",
        "get_union_prompt — similar guidance; no canonical hints for UNION",
        "Identical",
    ),
    # ── mcts_expand: Config Rules — GROUP_BY/AGGREGATE ────────────────────
    (
        "mcts_expand",
        "Config Rules — GROUP_BY/AGGREGATE",
        "[get_hints_section(GROUPBY_AGG_HINT_IDS, fmt='bullet')]\nHint IDs: [10,14,16,17,18,20,21]\n"
        "+ Format: GROUP_BY: [table.col1, ...], AGGREGATIONS: [COUNT(table.col), SUM(table.col2), ...]",
        "get_group_by_aggregate_prompt — get_hints_section(GROUPBY_AGG_HINT_IDS, fmt='bullet')",
        "Needs Update → Done",
    ),
    # ── mcts_expand: Config Rules — PIVOT/UNPIVOT, NO_MORE_OPERATION ──────
    (
        "mcts_expand",
        "Config Rules — PIVOT/UNPIVOT, NO_MORE_OPERATION",
        "PIVOT / UNPIVOT — no additional configuration needed. Format: (just the operator line)\n"
        "NO_MORE_OPERATION — the pipeline is complete. Format: (just the operator line)",
        "No direct multistep equivalent (multistep issues separate prompts per operator type)",
        "No Multistep Equivalent",
    ),
    # ── mcts_expand: Selection Guidance (Hints) ───────────────────────────
    (
        "mcts_expand",
        "Selection Guidance (Canonical Hints)",
        "[get_hints_section(NEXT_OPERATOR_HINT_IDS, fmt='bullet')]\nHint IDs: [1,2,3,4,5,6,9,11,16]",
        "get_next_operator_prompt — get_hints_section(NEXT_OPERATOR_HINT_IDS, fmt='bullet')",
        "Needs Update → Done",
    ),
    # ── mcts_expand: MCTS-specific rules ──────────────────────────────────
    (
        "mcts_expand",
        "Selection Guidance (MCTS-specific rules)",
        "- If the operation history already covers all needed transformations → propose NO_MORE_OPERATION.\n"
        "- Do NOT repeat an operation+configuration already present in the operation history.",
        "No multistep equivalent (MCTS tree-search specific)",
        "No Multistep Equivalent",
    ),
    # ── mcts_expand: Output Format ────────────────────────────────────────
    (
        "mcts_expand",
        "Output Format",
        "$CANDIDATE N$ ... $END$ delimited blocks with OPERATOR: <TYPE> + configuration lines. "
        "Up to k candidates ranked most-to-least promising.",
        "multistep uses $OPERATOR$ delimited single-answer format",
        "No Multistep Equivalent",
    ),
    # ── mcts_simulate: Task Description ───────────────────────────────────
    (
        "mcts_simulate",
        "Task Description",
        "You are generating executable Python code at runtime. Your goal is to convert multiple "
        "source tables into the format of the target table.",
        "get_python_script_simple — identical task description",
        "Identical",
    ),
    # ── mcts_simulate: Partial Plan section ───────────────────────────────
    (
        "mcts_simulate",
        "Partial Plan section (MCTS-specific)",
        "Partial Operation Plan (treat this as a STARTING GUIDE, not a fixed sequence). "
        "Use operations as a hint for first transformation step(s), then reason about additional "
        "steps required to produce the correct target schema.",
        "get_python_script_simple — Operation History used as strict sequence, not a guide",
        "No Multistep Equivalent",
    ),
    # ── mcts_simulate: Context fields ─────────────────────────────────────
    (
        "mcts_simulate",
        "Context fields",
        "Target Table Name, Target Schema, Target Examples, "
        "Source Information (with paths), CSV save path",
        "get_python_script_simple — same fields",
        "Identical",
    ),
    # ── mcts_simulate: Step-by-step thinking ──────────────────────────────
    (
        "mcts_simulate",
        "Step-by-step thinking (MCTS-specific)",
        "Before writing code, THINK STEP BY STEP about: (a) what partial plan implies for first "
        "operation(s), (b) intermediate schema produced, (c) further operations needed, "
        "(d) data-type fixes or renames required.",
        "No equivalent in get_python_script_simple",
        "No Multistep Equivalent",
    ),
    # ── mcts_simulate: Plan output block ──────────────────────────────────
    (
        "mcts_simulate",
        "Plan output block (MCTS-specific)",
        "$PLAN$ ... NO_MORE_OPERATION ... $END_PLAN$ block with one operation per line. "
        "Same format as operation history.",
        "No equivalent in get_python_script_simple",
        "No Multistep Equivalent",
    ),
    # ── mcts_simulate: Code generation instruction ────────────────────────
    (
        "mcts_simulate",
        "Code generation instruction",
        "Generate Python script implementing COMPLETE transformation from raw source tables to "
        "final target CSV. Immediately executable — no placeholders. Quote between ```Python and ```.",
        "get_python_script_simple — identical instruction",
        "Identical",
    ),
    # ── mcts_simulate: Hints ──────────────────────────────────────────────
    (
        "mcts_simulate",
        "Hints (under '══ Hints for Python code generation ══' header)",
        "[get_hints_section(PYTHON_SCRIPT_HINT_IDS, fmt='bullet')]\n"
        "Hint IDs: [1,2,3,4,5,10,11,16,17,24,27,28,29,30,31,32]\n"
        "Replaces previous: 14-bullet intro list + Hint 1-17 numbered block",
        "get_python_script_simple — get_hints_section(PYTHON_SCRIPT_HINT_IDS, fmt='bullet')",
        "Needs Update → Done",
    ),
    # ── mcts_simulate: Error feedback ─────────────────────────────────────
    (
        "mcts_simulate",
        "Error feedback",
        "Errors in previous attempts: {error_string}",
        "get_python_script_simple — same error feedback line",
        "Identical",
    ),
]


def build_lg_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet(title=LG_SHEET_NAME)

    headers = [
        "MCTS Component",
        "Section Type",
        "Current/Updated Text Summary",
        "Maps To (Multistep)",
        "Match Status",
    ]

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Write data rows
    for row_idx, row_data in enumerate(LG_ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Color the Match Status cell (column 5)
        status = row_data[4]
        if status in FILL_COLORS:
            ws.cell(row=row_idx, column=5).fill = PatternFill(
                "solid", fgColor=FILL_COLORS[status]
            )

    # Set column widths
    col_widths = [18, 36, 72, 60, 26]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    if LG_SHEET_NAME in wb.sheetnames:
        del wb[LG_SHEET_NAME]
    build_sheet(wb)
    build_lg_sheet(wb)
    wb.save(XLSX_PATH)
    print(f"Sheet '{SHEET_NAME}' added to {XLSX_PATH} ({len(ROWS)} rows).")
    print(f"Sheet '{LG_SHEET_NAME}' added to {XLSX_PATH} ({len(LG_ROWS)} rows).")


if __name__ == "__main__":
    main()
