"""
Generate a two-sheet Excel comparison for len_4 experiments (cases 15-99).

Sheet 1 - Method comparison: Multistep | Multistep+Critique | AgentFlow | Langraph
          Metrics: Accuracy, Avg Cost, Avg Latency (only over cases done in all methods)

Sheet 2 - Per-case breakdown: one row per case, one col per method
          Values: True / False / Remaining
"""

import csv
import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = "/home/local/ASUAD/jrtandel/transchema"

# ── File paths ────────────────────────────────────────────────────────────────

CRITIQUE_LOG = os.path.join(ROOT, "logs-auto-suggest-llm-21-04")
MULTISTEP_CSVS = [
    f"{CRITIQUE_LOG}/len_4_rerun_without_materialization_15_48_20260305_123232/results/multi_step.csv",
    f"{CRITIQUE_LOG}/len_4_rerun_without_materialization_48_81_20260305_123238/results/multi_step.csv",
    f"{CRITIQUE_LOG}/len_4_rerun_without_materialization_81_100_20260305_123252/results/multi_step.csv",
]
CRITIQUE_CSVS = [
    f"{CRITIQUE_LOG}/len_4_rerun_without_materialization_15_48_20260305_123232/results/critique.csv",
    f"{CRITIQUE_LOG}/len_4_rerun_without_materialization_48_81_20260305_123238/results/critique.csv",
    f"{CRITIQUE_LOG}/len_4_rerun_without_materialization_81_100_20260305_123252/results/critique.csv",
]
AGENTFLOW_CSVS = [
    f"{ROOT}/AgentFlow/results/len_4_15_48_20260305_123306/results_summary.csv",
    f"{ROOT}/AgentFlow/results/len_4_48_81_20260305_123322/results_summary.csv",
    f"{ROOT}/AgentFlow/results/len_4_81_100_20260305_123334/results_summary.csv",
]
LANGRAPH_CSVS = [
    f"{ROOT}/Langraph/results_langraph/len_4_15_47_20260305_123400/results_summary.csv",
    f"{ROOT}/Langraph/results_langraph/len_4_48_80_20260305_123412/results_summary.csv",
    f"{ROOT}/Langraph/results_langraph/len_4_81_99_20260305_123424/results_summary.csv",
]

ALL_CASES = [f"4_{i}" for i in range(15, 100)]  # 85 cases


# ── Parsers ───────────────────────────────────────────────────────────────────

def parse_bool(val):
    return str(val).strip().lower() == "true"

def safe_float(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def load_multistep(csv_paths):
    """Returns dict: case_id -> {correct, cost, latency}
    NOTE: The CSV has 7 header columns but only 6 data fields per row.
    Actual mapping: "Soft Match" = cost ($), "Soft Acc" = latency (s).
    The "Cost" column holds the soft score; "Latency" holds operation history.
    """
    data = {}
    for path in csv_paths:
        if not os.path.exists(path):
            continue
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                cid = row["Length"].strip()
                data[cid] = {
                    "correct": parse_bool(row["Hard Match"]),
                    "cost": safe_float(row["Soft Match"]),
                    "latency": safe_float(row["Soft Acc"]),
                }
    return data

def load_critique(csv_paths):
    """Returns dict: case_id -> {correct, cost, latency}
    Aggregates all critique types per case: correct=any True, cost/latency=sum.
    Same column misalignment as multi_step: "Soft Match" = cost, "Soft Acc" = latency.
    """
    raw = defaultdict(lambda: {"correct": False, "cost": 0.0, "latency": 0.0})
    for path in csv_paths:
        if not os.path.exists(path):
            continue
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                cid = row["Length"].strip()
                raw[cid]["correct"] = raw[cid]["correct"] or parse_bool(row["Hard Match"])
                raw[cid]["cost"] += safe_float(row["Soft Match"])
                raw[cid]["latency"] += safe_float(row["Soft Acc"])
    return dict(raw)

def load_agentflow(csv_paths):
    """Returns dict: case_id -> {correct, cost, latency}"""
    data = {}
    for path in csv_paths:
        if not os.path.exists(path):
            continue
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                cid = row["case_id"].strip()
                data[cid] = {
                    "correct": parse_bool(row["is_correct"]),
                    "cost": safe_float(row["cost"]),
                    "latency": safe_float(row["latency_seconds"]),
                }
    return data

def load_langraph(csv_paths):
    """Returns dict: case_id -> {correct, cost, latency}"""
    data = {}
    for path in csv_paths:
        if not os.path.exists(path):
            continue
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                cid = row["case_id"].strip()
                data[cid] = {
                    "correct": parse_bool(row["is_correct"]),
                    "cost": safe_float(row["cost"]),
                    "latency": safe_float(row["latency_seconds"]),
                }
    return data


# ── Build combined Multistep+Critique ────────────────────────────────────────

def combine_ms_critique(ms_data, crit_data):
    """
    Combined result: correct if either MS or Critique correct.
    Cost/latency = MS + Critique (Critique only run if MS failed, but we sum all rows logged).
    """
    combined = {}
    all_keys = set(ms_data) | set(crit_data)
    for cid in all_keys:
        ms = ms_data.get(cid)
        cr = crit_data.get(cid)
        if ms and cr:
            combined[cid] = {
                "correct": ms["correct"] or cr["correct"],
                "cost": ms["cost"] + cr["cost"],
                "latency": ms["latency"] + cr["latency"],
            }
        elif ms:
            combined[cid] = ms.copy()
        elif cr:
            combined[cid] = cr.copy()
    return combined


# ── Excel helpers ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="2F4F8F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TRUE_FILL  = PatternFill("solid", fgColor="C6EFCE")
FALSE_FILL = PatternFill("solid", fgColor="FFC7CE")
REM_FILL   = PatternFill("solid", fgColor="FFEB9C")
ALT_FILL   = PatternFill("solid", fgColor="EEF2FF")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, value):
    cell.value = value
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def val(cell, value, fill=None):
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
    if fill:
        cell.fill = fill


# ── Sheet 1: Method comparison ────────────────────────────────────────────────

def write_comparison_sheet(ws, ms, ms_crit, af, lg):
    methods = [
        ("Multistep",            ms),
        ("Multistep + Critique", ms_crit),
        ("AgentFlow",            af),
        ("Langraph (MCTS)",      lg),
    ]

    headers = ["Method", "Correct", "Total Cases",
               "Accuracy (%)", "Avg Cost ($)", "Avg Latency (s)"]
    for col, h in enumerate(headers, 1):
        hdr(ws.cell(1, col), h)

    ws.row_dimensions[1].height = 30
    for col, w in enumerate([22, 10, 13, 15, 14, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    for row_idx, (name, data) in enumerate(methods, 2):
        # Accuracy over all 85 cases (missing = failed/incorrect)
        correct  = sum(1 for c in ALL_CASES if data.get(c, {}).get("correct", False))
        accuracy = correct / len(ALL_CASES) * 100

        # Cost and latency only over cases that have data (exclude failed/missing)
        done_cases = [c for c in ALL_CASES if c in data]
        avg_cost = (sum(data[c]["cost"]    for c in done_cases) / len(done_cases)) if done_cases else 0
        avg_lat  = (sum(data[c]["latency"] for c in done_cases) / len(done_cases)) if done_cases else 0

        fill = ALT_FILL if row_idx % 2 == 0 else None

        def fval(cell, v, f=fill):
            cell.value = v
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            if f:
                cell.fill = f

        fval(ws.cell(row_idx, 1), name)
        ws.cell(row_idx, 1).font = Font(bold=True)
        fval(ws.cell(row_idx, 2), correct)
        fval(ws.cell(row_idx, 3), len(ALL_CASES))
        fval(ws.cell(row_idx, 4), round(accuracy, 1))
        fval(ws.cell(row_idx, 5), round(avg_cost, 4))
        fval(ws.cell(row_idx, 6), round(avg_lat, 1))

    note_row = len(methods) + 3
    ws.cell(note_row, 1).value = (
        "Accuracy over all 85 cases (missing cases counted as failed). "
        "Avg Cost and Latency computed only over cases with logged data."
    )
    ws.cell(note_row, 1).font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)


# ── Sheet 2: Per-case breakdown ───────────────────────────────────────────────

def write_percase_sheet(ws, ms, ms_crit, af, lg):
    methods = [
        ("Multistep",            ms),
        ("Multistep + Critique", ms_crit),
        ("AgentFlow",            af),
        ("Langraph (MCTS)",      lg),
    ]

    headers = ["Case"] + [m[0] for m in methods]
    for col, h in enumerate(headers, 1):
        hdr(ws.cell(1, col), h)

    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 10
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    for row_idx, case_id in enumerate(ALL_CASES, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None

        name_cell = ws.cell(row_idx, 1)
        name_cell.value = case_id
        name_cell.alignment = Alignment(horizontal="center", vertical="center")
        name_cell.border = BORDER
        if fill:
            name_cell.fill = fill

        for col_idx, (_, data) in enumerate(methods, 2):
            cell = ws.cell(row_idx, col_idx)
            if case_id not in data:
                val(cell, "Failed", FALSE_FILL)
            elif data[case_id]["correct"]:
                val(cell, "Correct", TRUE_FILL)
            else:
                val(cell, "Incorrect", FALSE_FILL)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading data...")
    ms_data   = load_multistep(MULTISTEP_CSVS)
    crit_data = load_critique(CRITIQUE_CSVS)
    af_data   = load_agentflow(AGENTFLOW_CSVS)
    lg_data   = load_langraph(LANGRAPH_CSVS)
    ms_crit   = combine_ms_critique(ms_data, crit_data)

    print(f"  Multistep:            {len(ms_data)} cases")
    print(f"  Multistep+Critique:   {len(ms_crit)} cases")
    print(f"  AgentFlow:            {len(af_data)} cases")
    print(f"  Langraph:             {len(lg_data)} cases")

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Method Comparison"
    ws2 = wb.create_sheet("Per-Case Breakdown")

    write_comparison_sheet(ws1, ms_data, ms_crit, af_data, lg_data)
    write_percase_sheet(ws2, ms_data, ms_crit, af_data, lg_data)

    out = os.path.join(ROOT, "results_len4_comparison.xlsx")
    wb.save(out)
    print(f"\nSaved: {out}")


if __name__ == "__main__":
    main()
