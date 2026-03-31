"""
Rebuilds results_memory_granularity.xlsx with full L1+L4 run data.
Sheets:
  - Summary Stats   : aggregate accuracy per mode
  - Per-Case (L1)   : all L1 cases (0-99) across summary/response/qa
  - Per-Case (L4)   : all L4 cases (18-99) across summary/response/qa
  - Token Limit     : cases that hit token limit per mode
  - Sources         : log dirs, CSV paths, python commands
"""

import csv, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "logs-auto-suggest-llm-21-04"

# ── Run directories ───────────────────────────────────────────────────────────
RUNS = {
    "summary_pilot": {
        "label": "Summary (pilot L1 0-9)",
        "dir":   "memory_granularity_summary_20260325_141429_v2_20260325_142209",
        "cases": [f"1_{i}" for i in range(10)],
    },
    "response_full": {
        "label": "Response (full L1+L4)",
        "dir":   "mem_gran_response_20260325_145000_20260325_145005",
        "cases": [f"1_{i}" for i in range(100)] + [f"4_{i}" for i in range(18,100)],
    },
    "qa_full": {
        "label": "QA (full L1+L4)",
        "dir":   "mem_gran_qa_20260325_145000_20260325_185628",
        "cases": [f"1_{i}" for i in range(100)] + [f"4_{i}" for i in range(18,100)],
    },
}

EXPECTED_L1 = [f"1_{i}" for i in range(100)]
EXPECTED_L4 = [f"4_{i}" for i in range(18, 100)]
ALL_EXPECTED = EXPECTED_L1 + EXPECTED_L4

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="BDD7EE")
TRUE_FILL = PatternFill("solid", fgColor="C6EFCE")
TRUE_FONT = Font(color="375623")
FALS_FILL = PatternFill("solid", fgColor="FFC7CE")
FALS_FONT = Font(color="9C0006")
PEND_FILL = PatternFill("solid", fgColor="FFEB9C")
PEND_FONT = Font(color="9C5700")
TL_FILL   = PatternFill("solid", fgColor="EDEDED")
TL_FONT   = Font(color="808080")
BOLD      = Font(bold=True)
CTR       = Alignment(horizontal="center", vertical="center")
WRAP_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = BOLD; c.fill = HDR_FILL
    c.alignment = WRAP_CTR; c.border = thin()
    return c

def vcell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.border = thin(); c.alignment = CTR
    if value is True:
        c.fill = TRUE_FILL; c.font = TRUE_FONT
    elif value is False:
        c.fill = FALS_FILL; c.font = FALS_FONT
    elif value == "TOKEN_LIMIT":
        c.fill = TL_FILL; c.font = TL_FONT
    elif value == "PENDING":
        c.fill = PEND_FILL; c.font = PEND_FONT
    return c

# ── Data loading ──────────────────────────────────────────────────────────────
def load_run(run_key):
    info = RUNS[run_key]
    exp_dir = info["dir"]
    ms_path   = f"{BASE}/{exp_dir}/results/average_multi_step.csv"
    crit_path = f"{BASE}/{exp_dir}/results/critique.csv"

    ms, costs, lats = {}, [], []
    try:
        with open(ms_path) as f:
            for row in csv.DictReader(f):
                ms[row["Length"]] = row["Hard Match"].strip() == "True"
                try: costs.append(float(row["Soft Match"]))
                except: pass
                try: lats.append(float(row["Cost"]))
                except: pass
    except FileNotFoundError:
        pass

    crit = {}
    try:
        with open(crit_path) as f:
            for row in csv.DictReader(f):
                cid = row["Length"]
                val = row["Hard Match"].strip() == "True"
                crit[cid] = crit.get(cid, False) or val
    except FileNotFoundError:
        pass

    expected = set(info["cases"])
    ran = set(ms.keys())
    token_limit = sorted(expected - ran,
                         key=lambda x: (int(x.split("_")[0]), int(x.split("_")[1])))

    return {
        "ms": ms, "crit": crit,
        "token_limit": token_limit,
        "avg_cost": sum(costs)/len(costs) if costs else 0,
        "avg_lat":  sum(lats)/len(lats) if lats else 0,
    }

def stats(ms, crit):
    ms_t  = {c for c,v in ms.items() if v}
    cr_t  = {c for c,v in crit.items() if v}
    corr  = ms_t | cr_t
    n     = len(ms)
    return {
        "n": n, "ms_pass": len(ms_t),
        "crit_only": len(cr_t - ms_t),
        "overlap": len(ms_t & cr_t),
        "correct": len(corr),
        "acc": 100*len(corr)/n if n else 0,
    }

data = {k: load_run(k) for k in RUNS}

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: Summary Stats
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Summary Stats")

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
t = ws.cell(row=1, column=1, value="Memory Granularity — Aggregate Statistics")
t.font = Font(bold=True, size=13); t.alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 22

SCOLS = ["Mode", "Length", "Cases Run", "Token Limit (skipped)",
         "MS Pass", "MS Pass%",
         "Critique-Only Pass", "Overlap",
         "Total Correct", "Accuracy%",
         "Avg Cost ($)", "Avg Latency (s)"]
for c, h in enumerate(SCOLS, 1):
    hdr(ws, 2, c, h)
ws.row_dimensions[2].height = 30

def stats_for_prefix(ms, crit, token_limit, prefix, costs_lats):
    ms_f  = {c: v for c, v in ms.items()   if c.startswith(prefix)}
    cr_f  = {c: v for c, v in crit.items() if c.startswith(prefix)}
    tl_f  = [c for c in token_limit         if c.startswith(prefix)]
    s     = stats(ms_f, cr_f)
    cl    = [(co, la) for cid, (co, la) in costs_lats.items() if cid.startswith(prefix)]
    avg_cost = sum(x[0] for x in cl) / len(cl) if cl else 0
    avg_lat  = sum(x[1] for x in cl) / len(cl) if cl else 0
    return s, len(tl_f), avg_cost, avg_lat

def case_cost_lat_dict(run_key):
    exp_dir = RUNS[run_key]["dir"]
    path = f"{BASE}/{exp_dir}/results/average_multi_step.csv"
    result = {}
    try:
        with open(path) as f:
            for row in csv.DictReader(f):
                try:
                    result[row["Length"]] = (float(row["Soft Match"]), float(row["Cost"]))
                except: pass
    except FileNotFoundError: pass
    return result

# (run_key, length_label, prefix)
rows_data = [
    ("summary_pilot", "L1 (0–9, pilot)", "1_"),
    ("response_full", "L1 (0–99)",        "1_"),
    ("response_full", "L4 (18–99)",       "4_"),
    ("qa_full",       "L1 (0–99)",        "1_"),
    ("qa_full",       "L4 (18–99)",       "4_"),
]

# Group label rows: shade alternating modes
MODE_FILLS = {
    "summary_pilot": PatternFill("solid", fgColor="F2F2F2"),
    "response_full": PatternFill("solid", fgColor="EBF3FB"),
    "qa_full":       PatternFill("solid", fgColor="EBF1DE"),
}

r = 3
prev_key = None
for key, length_label, prefix in rows_data:
    d  = data[key]
    cl = case_cost_lat_dict(key)
    s, tl_count, avg_cost, avg_lat = stats_for_prefix(
        d["ms"], d["crit"], d["token_limit"], prefix, cl)

    mode_label = RUNS[key]["label"] if key != prev_key else ""
    prev_key = key

    vals = [
        mode_label, length_label,
        s["n"], tl_count,
        s["ms_pass"], f"{100*s['ms_pass']/s['n']:.1f}%" if s["n"] else "–",
        s["crit_only"], s["overlap"],
        f"{s['correct']}/{s['n']}", f"{s['acc']:.1f}%",
        round(avg_cost, 5), round(avg_lat, 2),
    ]
    row_fill = MODE_FILLS[key]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = thin(); cell.alignment = CTR
        cell.fill = row_fill
        if c == 4 and tl_count > 0:
            cell.fill = TL_FILL; cell.font = TL_FONT
        if c == 1 and mode_label:
            cell.font = BOLD
    r += 1

ws.freeze_panes = "A3"
for i, w in enumerate([26,16,10,20,9,9,18,9,14,11,13,15], 1):
    ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# Sheets 2-3: Per-Case (L1) and Per-Case (L4)
# ══════════════════════════════════════════════════════════════════════════════
# Columns: Case ID | Summary MS | Summary Crit | Response MS | Response Crit | QA MS | QA Crit | [spacer] | Resp Cost | Resp Lat | QA Cost | QA Lat
#  Summary pilot only has L1 0-9, so L1 10-99 and all L4 show None for summary

for sheet_name, case_list in [("Per-Case (L1)", EXPECTED_L1), ("Per-Case (L4)", EXPECTED_L4)]:
    ws = wb.create_sheet(sheet_name)

    length_label = "L1 (0–99)" if "L1" in sheet_name else "L4 (18–99)"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    t = ws.cell(row=1, column=1,
        value=f"Memory Granularity Per-Case — {length_label} (gpt-4.1-mini, GT judge)")
    t.font = Font(bold=True, size=12); t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20

    # Group headers row 2
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    g1 = ws.cell(row=2, column=2, value="Pass / Fail")
    g1.font = BOLD; g1.fill = PatternFill("solid", fgColor="D9E1F2")
    g1.alignment = Alignment(horizontal="center"); g1.border = thin()

    ws.cell(row=2, column=8).border = thin()  # spacer

    ws.merge_cells(start_row=2, start_column=9, end_row=2, end_column=12)
    g2 = ws.cell(row=2, column=9, value="Cost & Latency")
    g2.font = BOLD; g2.fill = PatternFill("solid", fgColor="E2EFDA")
    g2.alignment = Alignment(horizontal="center"); g2.border = thin()

    # Column headers row 3
    for c, h in enumerate([
        "Case ID",
        "Summary\nMS", "Summary\nCritique",
        "Response\nMS", "Response\nCritique",
        "QA\nMS", "QA\nCritique",
        "",
        "Response\nCost ($)", "Response\nLat (s)",
        "QA\nCost ($)", "QA\nLat (s)",
    ], 1):
        hdr(ws, 3, c, h)
    ws.row_dimensions[3].height = 40

    sum_ms   = data["summary_pilot"]["ms"]
    sum_crit = data["summary_pilot"]["crit"]
    resp_ms  = data["response_full"]["ms"]
    resp_crit= data["response_full"]["crit"]
    qa_ms    = data["qa_full"]["ms"]
    qa_crit  = data["qa_full"]["crit"]
    resp_tl  = set(data["response_full"]["token_limit"])
    qa_tl    = set(data["qa_full"]["token_limit"])

    # Load per-case cost/lat from response and qa
    def case_cost_lat(run_key):
        exp_dir = RUNS[run_key]["dir"]
        path = f"{BASE}/{exp_dir}/results/average_multi_step.csv"
        result = {}
        try:
            with open(path) as f:
                for row in csv.DictReader(f):
                    try:
                        result[row["Length"]] = (
                            float(row["Soft Match"]),  # cost $
                            float(row["Cost"])          # latency s
                        )
                    except:
                        pass
        except FileNotFoundError:
            pass
        return result

    resp_cl = case_cost_lat("response_full")
    qa_cl   = case_cost_lat("qa_full")

    for row_idx, case_id in enumerate(case_list, start=4):
        vcell(ws, row_idx, 1, case_id)

        # Summary: only pilot cases have data
        if case_id in sum_ms:
            vcell(ws, row_idx, 2, sum_ms[case_id])
            vcell(ws, row_idx, 3, sum_crit.get(case_id, None))
        else:
            vcell(ws, row_idx, 2, None)
            vcell(ws, row_idx, 3, None)

        # Response
        if case_id in resp_tl:
            vcell(ws, row_idx, 4, "TOKEN_LIMIT")
            vcell(ws, row_idx, 5, "TOKEN_LIMIT")
        elif case_id in resp_ms:
            vcell(ws, row_idx, 4, resp_ms[case_id])
            vcell(ws, row_idx, 5, resp_crit.get(case_id, None))
        else:
            vcell(ws, row_idx, 4, None)
            vcell(ws, row_idx, 5, None)

        # QA
        if case_id in qa_tl:
            vcell(ws, row_idx, 6, "TOKEN_LIMIT")
            vcell(ws, row_idx, 7, "TOKEN_LIMIT")
        elif case_id in qa_ms:
            vcell(ws, row_idx, 6, qa_ms[case_id])
            vcell(ws, row_idx, 7, qa_crit.get(case_id, None))
        else:
            vcell(ws, row_idx, 6, None)
            vcell(ws, row_idx, 7, None)

        # Spacer
        ws.cell(row=row_idx, column=8).border = thin()

        # Cost/lat
        for col, cl_dict, tl_set in [(9, resp_cl, resp_tl), (11, qa_cl, qa_tl)]:
            if case_id in tl_set:
                vcell(ws, row_idx, col,   "TOKEN_LIMIT")
                vcell(ws, row_idx, col+1, "TOKEN_LIMIT")
            elif case_id in cl_dict:
                cost, lat = cl_dict[case_id]
                c1 = ws.cell(row=row_idx, column=col,   value=round(cost,5))
                c2 = ws.cell(row=row_idx, column=col+1, value=round(lat,2))
                c1.border = thin(); c1.alignment = CTR
                c2.border = thin(); c2.alignment = CTR
            else:
                vcell(ws, row_idx, col,   None)
                vcell(ws, row_idx, col+1, None)

    ws.freeze_panes = "B4"
    for i, w in enumerate([9, 11,11, 13,13, 9,9, 3, 14,12, 14,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    print(f"✓ {sheet_name} built ({len(case_list)} cases)")


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4: Token Limit Cases
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Token Limit Cases")

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
t = ws.cell(row=1, column=1, value="Cases That Hit Token Limit (skipped — no result produced)")
t.font = Font(bold=True, size=12); t.alignment = Alignment(horizontal="center")

for c, h in enumerate(["Mode", "Length Group", "Count", "Case IDs"], 1):
    hdr(ws, 2, c, h)

r = 3
for key in ["summary_pilot", "response_full", "qa_full"]:
    tl = data[key]["token_limit"]
    l1_tl = [c for c in tl if c.startswith("1_")]
    l4_tl = [c for c in tl if c.startswith("4_")]
    label = RUNS[key]["label"]
    for group, group_tl in [("L1", l1_tl), ("L4 (18+)", l4_tl)]:
        if not group_tl and key == "summary_pilot":
            continue
        cell_ids = ", ".join(group_tl) if group_tl else "none"
        for c, v in enumerate([label, group, len(group_tl), cell_ids], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = thin()
            cell.alignment = Alignment(vertical="top", wrap_text=(c==4))
            if len(group_tl) > 0 and c == 3:
                cell.fill = TL_FILL; cell.font = TL_FONT
        ws.row_dimensions[r].height = max(15, min(15 * ((len(cell_ids)//80)+1), 60))
        r += 1

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 8
ws.column_dimensions["D"].width = 100
ws.freeze_panes = "A3"
print("✓ Token Limit Cases sheet built")


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5: Sources
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Sources")

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
t = ws.cell(row=1, column=1, value="Source Paths & Commands — Memory Granularity Experiments")
t.font = Font(bold=True, size=13); t.alignment = Alignment(horizontal="center")

SCOLS2 = ["Mode", "Scope", "Status",
          "Experiment Name", "Log Directory",
          "average_multi_step.csv", "critique.csv", "final_critique.csv",
          "Python Command"]
for c, h in enumerate(SCOLS2, 1):
    hdr(ws, 2, c, h)

sources = [
    ("summary_pilot", "L1 0-9 (pilot)", False,
     "python3 critique_data.py \\\n"
     "  --model gpt-4.1-mini --judge gt \\\n"
     "  --cases 1_0 1_1 1_2 1_3 1_4 1_5 1_6 1_7 1_8 1_9 \\\n"
     "  --memory_granularity summary \\\n"
     "  --experiment-name mem_gran_summary_<timestamp>"),
    ("response_full", "L1 0-99 + L4 18-99", False,
     "# See run_mem_granularity.sh\n"
     "python3 critique_data.py \\\n"
     "  --model gpt-4.1-mini --judge gt \\\n"
     "  --cases $L1_CASES $L4_CASES \\\n"
     "  --memory_granularity response \\\n"
     "  --experiment-name mem_gran_response_$TS"),
    ("qa_full", "L1 0-99 + L4 18-99", False,
     "# See run_mem_granularity.sh\n"
     "python3 critique_data.py \\\n"
     "  --model gpt-4.1-mini --judge gt \\\n"
     "  --cases $L1_CASES $L4_CASES \\\n"
     "  --memory_granularity qa \\\n"
     "  --experiment-name mem_gran_qa_$TS"),
]

for r, (key, scope, pending, cmd) in enumerate(sources, start=3):
    exp_dir = RUNS[key]["dir"]
    log_dir = f"{BASE}/{exp_dir}"
    exp_name = exp_dir.rsplit("_202", 1)[0]
    vals = [
        RUNS[key]["label"], scope, "COMPLETE",
        exp_name, log_dir,
        f"{log_dir}/results/average_multi_step.csv",
        f"{log_dir}/results/critique.csv",
        f"{log_dir}/results/final_critique.csv",
        cmd,
    ]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = thin()
        cell.alignment = Alignment(vertical="top", wrap_text=(c==9))
    ws.row_dimensions[r].height = 85

ws.freeze_panes = "A3"
for i, w in enumerate([26,24,10,38,65,70,70,70,75], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
print("✓ Sources sheet built")


wb.save("results_memory_granularity.xlsx")
print("\n✓ Saved results_memory_granularity.xlsx")
