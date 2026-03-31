"""
Rebuilds results_gpt41_mini_o4_mini.xlsx:
  - Correct column mapping: CSV "Cost" = latency(s), CSV "Soft Match" = cost($)
  - Fixes Overall sheet cost/latency
  - Rebuilds L1, L4, L9 per-case sheets with all 8 model/judge columns
  - Per-case: MS result + Critique result, colour-coded
  - Deduplication: no case counted twice in accuracy
"""

import csv, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "logs-auto-suggest-llm-21-04"

# (label_in_sheet, model_str, judge_str, length, exp_dir)
RUNS = [
    ("gpt41_mini_gt",     "gpt-4.1-mini", "gt",              "L1", "judge_exp_gt_gpt41_mini_L1_20260325_135133"),
    ("gpt41_mini_hybrid", "gpt-4.1-mini", "llm_score_hybrid","L1", "judge_exp_llm_score_hybrid_gpt41_mini_L1_20260325_145701"),
    ("gpt41_mini_gt",     "gpt-4.1-mini", "gt",              "L4", "judge_exp_gt_gpt41_mini_L4_20260325_160016"),
    ("gpt41_mini_hybrid", "gpt-4.1-mini", "llm_score_hybrid","L4", "judge_exp_llm_score_hybrid_gpt41_mini_L4_20260325_184305"),
    ("gpt41_mini_gt",     "gpt-4.1-mini", "gt",              "L9", "judge_exp_gt_gpt41_mini_L9_20260325_215200"),
    ("gpt41_mini_hybrid", "gpt-4.1-mini", "llm_score_hybrid","L9", "judge_exp_llm_score_hybrid_gpt41_mini_L9_20260326_012146"),
    ("o4_mini_gt",        "o4-mini",      "gt",              "L9", "judge_exp_gt_o4_mini_L9_20260325_135041"),
    ("o4_mini_hybrid",    "o4-mini",      "llm_score_hybrid","L9", "judge_exp_llm_score_hybrid_o4_mini_L9_20260325_175242"),
    ("o4_mini_gt",        "o4-mini",      "gt",              "L4", "judge_exp_gt_o4_mini_L4_20260326_160550"),
    ("o4_mini_hybrid",    "o4-mini",      "llm_score_hybrid","L4", "judge_exp_llm_score_hybrid_o4_mini_L4_20260326_201625"),
    ("o4_mini_gt",        "o4-mini",      "gt",              "L1", "judge_exp_gt_o4_mini_L1_20260327_003643"),
    ("o4_mini_hybrid",    "o4-mini",      "llm_score_hybrid","L1", "judge_exp_llm_score_hybrid_o4_mini_L1_20260327_020504"),
]

PENDING = []

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="BDD7EE")
TRUE_FILL = PatternFill("solid", fgColor="C6EFCE")
TRUE_FONT = Font(color="375623")
FALS_FILL = PatternFill("solid", fgColor="FFC7CE")
FALS_FONT = Font(color="9C0006")
PEND_FILL = PatternFill("solid", fgColor="FFEB9C")
PEND_FONT = Font(color="9C5700")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
WARN_FONT = Font(color="C00000")
BOLD      = Font(bold=True)
CTR       = Alignment(horizontal="center", vertical="center")
WRAP_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, val, width=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = BOLD; c.fill = HDR_FILL; c.alignment = WRAP_CTR; c.border = thin()
    return c

def val_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.border = thin(); c.alignment = CTR
    if value is True:
        c.fill = TRUE_FILL; c.font = TRUE_FONT
    elif value is False:
        c.fill = FALS_FILL; c.font = FALS_FONT
    elif value == "PENDING":
        c.fill = PEND_FILL; c.font = PEND_FONT
    elif value in (None, ""):
        pass
    return c

# ── Data loading ──────────────────────────────────────────────────────────────
def load_run(exp_dir):
    """Returns (ms_dict, crit_dict, avg_cost, avg_lat)
    ms_dict:   case_id -> bool (GT hard match)
    crit_dict: case_id -> bool (best critique result across fd/metadata)
    avg_cost:  mean cost in $ (CSV "Soft Match" column)
    avg_lat:   mean latency in s (CSV "Cost" column)
    """
    ms_path   = f"{BASE}/{exp_dir}/results/average_multi_step.csv"
    crit_path = f"{BASE}/{exp_dir}/results/critique.csv"

    ms = {}
    costs, lats = [], []
    with open(ms_path) as f:
        for row in csv.DictReader(f):
            ms[row["Length"]] = row["Hard Match"].strip() == "True"
            try: costs.append(float(row["Soft Match"]))   # actual $
            except: pass
            try: lats.append(float(row["Cost"]))          # actual seconds
            except: pass

    crit = {}
    try:
        with open(crit_path) as f:
            for row in csv.DictReader(f):
                cid = row["Length"]
                val = row["Hard Match"].strip() == "True"
                crit[cid] = crit.get(cid, False) or val
    except FileNotFoundError:
        pass

    avg_cost = sum(costs)/len(costs) if costs else 0
    avg_lat  = sum(lats)/len(lats) if lats else 0
    return ms, crit, avg_cost, avg_lat


def compute_stats(ms, crit):
    ms_true   = {c for c,v in ms.items() if v}
    crit_true = {c for c,v in crit.items() if v}
    overlap   = ms_true & crit_true
    correct   = ms_true | crit_true
    total     = len(ms)
    return {
        "total": total,
        "ms_pass": len(ms_true),
        "crit_only": len(crit_true - ms_true),
        "overlap": len(overlap),
        "correct": len(correct),
        "acc_pct": 100*len(correct)/total if total else 0,
    }


# ── Load all run data ─────────────────────────────────────────────────────────
run_data = {}   # (label, length) -> (ms, crit, avg_cost, avg_lat)
for label, model, judge, length, exp_dir in RUNS:
    run_data[(label, length)] = load_run(exp_dir)


# ── Open workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook("results_gpt41_mini_o4_mini.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# 1. Rebuild Overall sheet
# ══════════════════════════════════════════════════════════════════════════════
del wb["Overall"]
ws = wb.create_sheet("Overall", 0)

OVERALL_COLS = [
    "Model", "Judge", "Length", "Cases Run",
    "MS Pass", "MS Pass%",
    "Critique-Only Pass",
    "Overlap (counted once)",
    "Total Correct (union)", "Accuracy%",
    "Avg Cost ($)", "Avg Latency (s)",
    "Note"
]

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(OVERALL_COLS))
t = ws.cell(row=1, column=1,
    value="GPT-4.1-mini & o4-mini — Overall Summary (Deduplicated, Corrected Cost/Latency)")
t.font = Font(bold=True, size=13)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 22

for c, h in enumerate(OVERALL_COLS, 1):
    hdr(ws, 2, c, h)
ws.row_dimensions[2].height = 32

r = 3
for label, model, judge, length, exp_dir in RUNS:
    ms, crit, avg_cost, avg_lat = run_data[(label, length)]
    s = compute_stats(ms, crit)
    has_overlap = s["overlap"] > 0
    note = f"⚠ {s['overlap']} cases in both MS+Critique" if has_overlap else "✓ No overlap"

    vals = [model, judge, length, s["total"],
            s["ms_pass"], f"{100*s['ms_pass']/s['total']:.1f}%",
            s["crit_only"], s["overlap"],
            f"{s['correct']}/{s['total']}", f"{s['acc_pct']:.1f}%",
            round(avg_cost, 5), round(avg_lat, 2), note]

    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = thin(); cell.alignment = CTR
        if has_overlap and c == 8:
            cell.fill = WARN_FILL; cell.font = WARN_FONT
    r += 1

for label, model, judge, length in PENDING:
    vals = [model, judge, length, "PENDING","","","","","","","","","Not yet run"]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = PEND_FILL; cell.font = PEND_FONT
        cell.border = thin(); cell.alignment = CTR
    r += 1

ws.freeze_panes = "A3"
col_widths = [16,20,8,10,9,9,19,20,16,11,13,16,30]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
print("✓ Overall sheet rebuilt")


# ══════════════════════════════════════════════════════════════════════════════
# 2. Rebuild L1 / L4 / L9 per-case sheets
# ══════════════════════════════════════════════════════════════════════════════
# Column layout per sheet:
#   Case ID |
#   gpt41_mini GT MS | gpt41_mini GT Crit |
#   gpt41_mini Hybrid MS | gpt41_mini Hybrid Crit |
#   o4_mini GT MS | o4_mini GT Crit |
#   o4_mini Hybrid MS | o4_mini Hybrid Crit |
#   [blank spacer]
#   MS Cost gpt41 GT ($) | MS Lat gpt41 GT (s) |
#   MS Cost gpt41 Hybrid ($) | MS Lat gpt41 Hybrid (s) |
#   MS Cost o4 GT ($) | MS Lat o4 GT (s) |
#   MS Cost o4 Hybrid ($) | MS Lat o4 Hybrid (s)

# per-case cost/lat from MS CSV
def load_case_cost_lat(exp_dir):
    path = f"{BASE}/{exp_dir}/results/average_multi_step.csv"
    result = {}
    try:
        with open(path) as f:
            for row in csv.DictReader(f):
                try:
                    cost = float(row["Soft Match"])   # $ cost
                    lat  = float(row["Cost"])         # seconds
                    result[row["Length"]] = (cost, lat)
                except:
                    result[row["Length"]] = (None, None)
    except FileNotFoundError:
        pass
    return result

LENGTH_RUNS = {
    "L1": [
        ("gpt41_mini_gt",     "gpt-4.1-mini", "GT",     "judge_exp_gt_gpt41_mini_L1_20260325_135133"),
        ("gpt41_mini_hybrid", "gpt-4.1-mini", "Hybrid", "judge_exp_llm_score_hybrid_gpt41_mini_L1_20260325_145701"),
        ("o4_mini_gt",        "o4-mini",      "GT",     "judge_exp_gt_o4_mini_L1_20260327_003643"),
        ("o4_mini_hybrid",    "o4-mini",      "Hybrid", "judge_exp_llm_score_hybrid_o4_mini_L1_20260327_020504"),
    ],
    "L4": [
        ("gpt41_mini_gt",     "gpt-4.1-mini", "GT",     "judge_exp_gt_gpt41_mini_L4_20260325_160016"),
        ("gpt41_mini_hybrid", "gpt-4.1-mini", "Hybrid", "judge_exp_llm_score_hybrid_gpt41_mini_L4_20260325_184305"),
        ("o4_mini_gt",        "o4-mini",      "GT",     "judge_exp_gt_o4_mini_L4_20260326_160550"),
        ("o4_mini_hybrid",    "o4-mini",      "Hybrid", "judge_exp_llm_score_hybrid_o4_mini_L4_20260326_201625"),
    ],
    "L9": [
        ("gpt41_mini_gt",     "gpt-4.1-mini", "GT",     "judge_exp_gt_gpt41_mini_L9_20260325_215200"),
        ("gpt41_mini_hybrid", "gpt-4.1-mini", "Hybrid", "judge_exp_llm_score_hybrid_gpt41_mini_L9_20260326_012146"),
        ("o4_mini_gt",        "o4-mini",      "GT",     "judge_exp_gt_o4_mini_L9_20260325_135041"),
        ("o4_mini_hybrid",    "o4-mini",      "Hybrid", "judge_exp_llm_score_hybrid_o4_mini_L9_20260325_175242"),
    ],
}

# Expected case IDs per length
def all_cases(length):
    L = length.replace("L","")
    start = 15 if L == "4" else 0
    return [f"{L}_{i}" for i in range(start, 100)]

for sheet_name, length_runs in LENGTH_RUNS.items():
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Load data for this length
    run_info = []   # (label, model, judge_str, ms_dict, crit_dict, cost_lat_dict)
    for label, model, judge_str, exp_dir in length_runs:
        if exp_dir:
            ms, crit, _, _ = load_run(exp_dir)
            cl = load_case_cost_lat(exp_dir)
        else:
            ms, crit, cl = {}, {}, {}
        run_info.append((label, model, judge_str, ms, crit, cl))

    cases = all_cases(sheet_name)

    # ── Build header rows (rows 1-3) ─────────────────────────────────────────
    # Row 1: Title
    total_cols = 1 + len(run_info)*2 + 1 + len(run_info)*2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(row=1, column=1,
        value=f"{sheet_name} Per-Case Breakdown — gpt-4.1-mini & o4-mini (GT + Hybrid judges)")
    t.font = Font(bold=True, size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Row 2: group headers
    # Results group
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1+len(run_info)*2)
    c2 = ws.cell(row=2, column=2, value="Pass / Fail Results")
    c2.font = BOLD; c2.fill = PatternFill("solid", fgColor="D9E1F2")
    c2.alignment = Alignment(horizontal="center"); c2.border = thin()
    # Cost/Latency group
    cost_col_start = 1 + len(run_info)*2 + 2
    ws.merge_cells(start_row=2, start_column=cost_col_start,
                   end_row=2, end_column=total_cols)
    c3 = ws.cell(row=2, column=cost_col_start, value="Cost & Latency (MS only)")
    c3.font = BOLD; c3.fill = PatternFill("solid", fgColor="E2EFDA")
    c3.alignment = Alignment(horizontal="center"); c3.border = thin()

    # Row 3: column headers
    hdr(ws, 3, 1, "Case ID")
    col = 2
    for label, model, judge_str, *_ in run_info:
        short = f"{model}\n{judge_str}"
        hdr(ws, 3, col,   f"MS\n{short}")
        hdr(ws, 3, col+1, f"Critique\n{short}")
        col += 2
    # spacer
    ws.cell(row=3, column=col, value="").border = thin()
    col += 1
    for label, model, judge_str, *_ in run_info:
        short = f"{model}\n{judge_str}"
        hdr(ws, 3, col,   f"Cost($)\n{short}")
        hdr(ws, 3, col+1, f"Lat(s)\n{short}")
        col += 2
    ws.row_dimensions[3].height = 46

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, case_id in enumerate(cases, start=4):
        val_cell(ws, row_idx, 1, case_id)
        col = 2
        for label, model, judge_str, ms, crit, cl in run_info:
            if not ms and not cl:  # pending run
                val_cell(ws, row_idx, col,   "PENDING")
                val_cell(ws, row_idx, col+1, "PENDING")
            else:
                ms_val   = ms.get(case_id, None)
                crit_val = crit.get(case_id, None)
                val_cell(ws, row_idx, col,   ms_val)
                val_cell(ws, row_idx, col+1, crit_val)
            col += 2
        # spacer
        ws.cell(row=row_idx, column=col).border = thin()
        col += 1
        for label, model, judge_str, ms, crit, cl in run_info:
            if cl:
                cost_v, lat_v = cl.get(case_id, (None, None))
                c1 = ws.cell(row=row_idx, column=col,   value=round(cost_v,5) if cost_v else None)
                c2 = ws.cell(row=row_idx, column=col+1, value=round(lat_v,2)  if lat_v  else None)
                c1.border = thin(); c1.alignment = CTR
                c2.border = thin(); c2.alignment = CTR
            else:
                val_cell(ws, row_idx, col,   "PENDING")
                val_cell(ws, row_idx, col+1, "PENDING")
            col += 2

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 9
    col = 2
    for _ in run_info:
        ws.column_dimensions[get_column_letter(col)].width   = 11
        ws.column_dimensions[get_column_letter(col+1)].width = 11
        col += 2
    ws.column_dimensions[get_column_letter(col)].width = 3   # spacer
    col += 1
    for _ in run_info:
        ws.column_dimensions[get_column_letter(col)].width   = 12
        ws.column_dimensions[get_column_letter(col+1)].width = 11
        col += 2

    ws.freeze_panes = "B4"
    print(f"✓ {sheet_name} sheet rebuilt ({len(cases)} cases, {len(run_info)} runs)")


wb.save("results_gpt41_mini_o4_mini.xlsx")
print("\n✓ Saved results_gpt41_mini_o4_mini.xlsx")
