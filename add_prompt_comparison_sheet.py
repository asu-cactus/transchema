"""
add_prompt_comparison_sheet.py
Adds a "Prompt Comparison" sheet to "Prompt Design.xlsx" comparing
Multi-Step, Agentflow, and LangGraph/MCTS prompts at operator granularity.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

XLSX_PATH = "/home/local/ASUAD/jrtandel/transchema/Prompt Design.xlsx"
SHEET_NAME = "Prompt Comparison"
PIPELINE_SHEET_NAME = "Pipeline Prompts"

# ── colour palette ─────────────────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # dark navy  – column headers
C_HEADER_FG   = "FFFFFF"
C_SECTION_BG  = "D6E4F7"   # light blue – step/operator section banner
C_MULTI_BG    = "FFF2CC"   # yellow     – Multi-Step rows
C_AGENT_BG    = "E2EFDA"   # green      – Agentflow rows
C_LG_BG       = "FCE4D6"   # peach      – LangGraph rows
C_NOTE_BG     = "F2F2F2"   # grey       – note rows

# ── raw prompt texts ───────────────────────────────────────────────────────

MULTI_OPERATOR_SELECTION = """\
query1
You are generating a data-pipeline to transform multiple source tables to target table and you need to answer "what operation should be performed next?". Take this decision based on "operation history", the schema of the source, target tables, and examples in the target table.
Allowed Operations: {allowed_operation_list}.
Operation History: {operation_history}

1. Target Table Name: {target_data_name}
2. Target Schema: {target_data_schema}
3. Target Examples: {target_samples}
4. Source Information: {source_information}

{fd_hints}

{hints[0]}

[Optional static hints from NEXT_OPERATOR_HINT_IDS]
- Please answer what operation you should perform next based on "operation history", "source tables" and "target tables" information ("schema" as well as the "examples") in one word.
- If you feel no more operation is needed further, please return 'NO_MORE_OPERATION'.
- You should only answer from allowed operations.
- Try not to repeat operation and it's configuration from the operation history.
- the final answer should be in $ quotes. i.e. $OPERATOR$\
"""

MULTI_JOIN = """\
query1
You are generating a data-pipeline to transform multiple source tables to target table and you need to answer "what tables should be joined and at which columns?". Take this decision based on "Operation History", "Source" and "Target" (table schema as well as the examples) information.

Allowed Operations: {allowed_operation_list}.

1. Target Table Name: {target_data_name}
2. Target Schema: {target_data_schema}
3. Target Examples: {target_samples}
4. Multi Source Information: {source_information}
5. Operation History: {operation_history}

-You may use the hint in your decision making process.
Hint : {hints[0]}
{fd_hints}

- Please answer which tables should be joined and at which columns that hasn't appeared yet in "operation history".
[Optional static hints from JOIN_HINT_IDS]
- You should only answer from available columns of the source tables.
- Only return two tables that should be joined and on which columns.
- Choose tables that contain columns similar to columns in the target tables to be joined.
- The final answer should be in format of list.
- The first element of list should be the list of two tables that should be joined.
- The next elements should be list of two columns on which the join should be perfomed.
- i.e. [ [table1, table2], [table1.join_column1,table2.join_column1], ...]
- The final answer list should be within $ quotes strictly. [i.e. $ Final Answer List $]\
"""

MULTI_UNION = """\
You are generating a data-pipeline to transform multiple source tables to target table and you need to answer "what tables should be Union-ed?". Take this decision based on "Operation History",few shot examples, "Source" and "Target" (table schema as well as the examples) information.

Allowed Operations: {allowed_operation_list}.

1. Target Table Name: {target_data_name}
2. Target Schema: {target_data_schema}
3. Target Examples: {target_samples}
4. Multi Source Information: {source_information}
5. Operation History: {operation_history}

- Please answer which tables should be unioned.
- Apply UNION to tables that have exactly the same schema without renaming columns.
- Choose tables having exactly the same schema to union.
- Try not to repeat operation and it's configuration. i.e. Union on tables should only appear once in "Operation History".
- You should only answer from available tables of the source tables.
- Please don't write your reasoning with the answer, just the answer would suffice.
- The final answer should be in format of list where elements are quoted by $. I.e. [$table1$, $table2$, ...].\
"""

MULTI_GROUPBY = """\
You are generating a data-pipeline to transform multiple source tables to target table and you need to answer "1. Which columns should be used for Group By operation? 2. Which columns should be Aggregated? 3.Which Aggregation functions should be used?". Take this decision based on "Operation History", few-shot-examples, "Source" and "Target" (table schema as well as the examples) information.

Allowed Operations: {allowed_operation_list}.

1. Target Table Name: {target_data_name}
2. Target Schema: {target_data_schema}
3. Target Examples: {target_samples}
4. Multi Source Information: {source_information}
5. Operation History: {operation_history}

-You may use the hint in your decision making process.
Hint : {hints[0]}
{fd_hints}

- Please answer on which columns "Group By" operation should be performed, on which columns aggregation should be performed and which aggregation functions should be used.
[Optional static hints from GROUPBY_AGG_HINT_IDS]
- You should only answer from available columns of the source tables.
- Please don't write your reasoning with the answer, just the answer would suffice.
- The final answer should be in the following format. lists where first list should be group by columns, next list should cover aggregation function and on which columns aggregations should be performed.
- Example : "group_by" = [table_name.group_by_column1, ...], "aggregations" = [agg_func(table_name.col), ...]\
"""

MULTI_PIVOT = "(No dedicated PIVOT prompt in Multi-Step; PIVOT is handled by the code generation prompt after selection)"

MULTI_UNPIVOT = "(No dedicated UNPIVOT prompt in Multi-Step; UNPIVOT is handled by the code generation prompt after selection)"

MULTI_CODE_GEN = """\
You are generating executable Python code at runtime. Please generate a Python script to convert multiple source tables to the format of the target table and STRICTLY follow the sequence of the operations mentioned in 'operation_history' list. The code should immediately executable in a correct way, which means it should NOT contain any placeholder for brievity. For example, even if there exists hundreds of source tables, these data needs to be loaded completely one by one or in a programmable way.

    Transformation Plan:

    Operation History: {operation_history}

    1. Target Table Name: {target_data_name}
    2. Target Schema: {target_data_schema}
    3. Target Examples: {target_samples}
    4. Source Information: {source_information_with_location}
    5. Write the result to this path {csv_save_path}

    Based on the transformation plan, generate the Python script that implements the transformation. The script should handle data import, transformation, and export. The script should be complete and executable, not omiting any single statement. For example, please list all the source paths that will be used.

    Please quote the Python script between one single "```Python" and "```".

[Optional hints from PYTHON_SCRIPT_HINT_IDS]
 Errors in previous Attempts : {error_string}\
"""

# ── Agentflow prompts ──────────────────────────────────────────────────────

AGENT_OPERATOR_SELECTION = """\
[PLANNER — _generate_next_step_operator()]

Task: Determine the optimal next step to build the data transformation pipeline operator by operator.

You are operating at OPERATOR granularity. You configure one operator at a time, incrementally building the pipeline.

CRITICAL CONSTRAINT: You can ONLY use tools from the available tools list.

=== OPERATOR TOOLS ===
- Configure_Join_Operator_Tool: Configure a JOIN between two tables.
- Configure_Union_Operator_Tool: Configure a UNION (stack) of tables with identical schemas.
- Configure_GroupBy_Aggregate_Operator_Tool: Configure GROUP BY + aggregation.
- Add_Pivot_Tool: Configure a PIVOT operation (wide format).
- Add_Unpivot_Tool: Configure an UNPIVOT (melt) operation (narrow format).

=== UTILITY TOOLS ===
- Code_Gen_And_Score_Tool: Generate Python code, execute it, and calculate score.
- Critique_Pipeline_Tool: Critique pipeline and suggest corrections.

=== DOMAIN HINTS FOR OPERATOR SELECTION ===
[Hints from NEXT_OPERATOR_HINT_IDS]

Context:
- Query: {question}
- Query Analysis: {query_analysis}
- Available Tools: {self.available_tools}
- Toolbox Metadata: {self.toolbox_metadata}
- Previous Steps and Their Results: {memory.get_actions()}
- Current Step: {step_count} of {max_step_count}
- Remaining Steps: {max_step_count - step_count}

Instructions:
1. Review the query, schemas, previous steps, and any score feedback.
2. Determine which operator to configure next, OR whether to generate/score code, OR whether to critique.
3. Select the single best tool for this step.
4. Formulate a specific, achievable sub-goal directly achievable by that tool.
5. Provide all necessary context.

Response Format:
1. Justification: Explain your reasoning and choice of tool.
2. Context: All information the tool needs.
3. Sub-Goal: The specific objective for the selected tool.
4. Tool Name: The exact name from the available tools list.\
"""

AGENT_JOIN = """\
[Configure_Join_Operator_Tool — get_join_prompt()]

You are generating a data-pipeline to transform multiple source tables to target table and you need to answer "what tables should be joined and at which columns?". Take this decision based on the transformation case information provided below.

{transformation_case_info}
(includes: Target Table Name, Target Schema, Target Examples, Multi Source Information, Operation History)

- Please answer which tables should be joined and at which columns that hasn't appeared yet in "operation history".
[Hints from JOIN_HINT_IDS]
- You should only answer from available columns of the source tables.
- Only return two tables that should be joined and on which columns.
- You may use the tables that are generated by the previous actions (e.g., join_result_0, join_result_1, etc.).
- Choose tables that contain columns similar to columns in the target tables to be joined.
- The final answer should be in format of list.
- The first element of list should be the list of two tables that should be joined.
- The next elements should be list of two columns on which the join should be perfomed.
- i.e. [ [table1, table2], [table1.join_column1,table2.join_column1], ...]
- The final answer list should be within $ quotes strictly. [i.e. $ Final Answer List $]\
"""

AGENT_UNION = """\
[Configure_Union_Operator_Tool — get_union_prompt()]

You are generating a data-pipeline to transform multiple source tables to target table and you need to answer "what tables should be Union-ed?". Take this decision based on "Operation History",few shot examples, "Source" and "Target" (table schema as well as the examples) information provided below.

{transformation_case_info}
(includes: Target Table Name, Target Schema, Target Examples, Multi Source Information, Operation History)

- Please answer which tables should be unioned.
- Apply UNION to tables that have exactly the same schema without renaming columns.
- Choose tables having exactly the same schema to union.
- Try not to repeat operation and it's configuration. i.e. Union on tables should only appear once in "Operation History".
- You should only answer from available tables of the source tables.
- Please don't write your reasoning with the answer, just the answer would suffice.
- The final answer should be in format of list where elements are quoted by $. I.e. [$table1$, $table2$, ...].\
"""

AGENT_GROUPBY = """\
[Configure_GroupBy_Aggregate_Operator_Tool — get_group_by_aggregate_prompt()]

You are generating a data-pipeline to transform multiple source tables to target table and you need to answer "1. Which columns should be used for Group By operation? 2. Which columns should be Aggregated? 3.Which Aggregation functions should be used?". Take this decision based on the transformation case information provided below.

{transformation_case_info}
(includes: Target Table Name, Target Schema, Target Examples, Multi Source Information, Operation History)

- Please answer on which columns "Group By" operation should be performed, on which columns aggregation should be performed and which aggregation functions should be used.
[Hints from GROUPBY_AGG_HINT_IDS]
- You should only answer from available columns of the source tables.
- Please don't write your reasoning with the answer, just the answer would suffice.
- The final answer should be in the following format:
  "group_by" = [table_name.group_by_column1, ...], "aggregations" = [agg_func(table_name.col), ...]\
"""

AGENT_PIVOT = """\
[Add_Pivot_Tool — get_pivot_prompt()]

You are configuring a PIVOT operation for a data pipeline that transforms multiple source tables into a target table.

A PIVOT operation takes rows and turns distinct values from one column into new column headers, aggregating values from another column.

{transformation_case_info}
(includes: Target Table Name, Target Schema, Target Examples, Source Information, Operation History)

Based on the source and target schema information above, determine the correct PIVOT configuration.

Instructions:
- Identify the "index" columns: columns kept as row identifiers.
- Identify the "columns" field: the source column whose distinct values become new column headers.
- Identify the "values" field: the source column whose values fill the pivoted cells.
- Identify the "aggfunc": aggregation function ("first", "sum", "mean", "count"). Use "first" if values are unique per cell.
- The result column names after pivot should match the target schema.
- You may reference intermediate results from the operation history (e.g., join_result_0).

Return ONLY the JSON pivot specification enclosed in $ signs:
${"index": ["col1", "col2"], "columns": "pivot_key_col", "values": "value_col", "aggfunc": "first"}$\
"""

AGENT_UNPIVOT = """\
[Add_Unpivot_Tool — get_unpivot_prompt()]

You are configuring an UNPIVOT (melt) operation for a data pipeline that transforms multiple source tables into a target table.

An UNPIVOT operation takes a wide table with multiple value columns and stacks those columns into rows, producing a narrower table with a key column and a value column.

{transformation_case_info}
(includes: Target Table Name, Target Schema, Target Examples, Source Information, Operation History)

Instructions:
- Identify "id_vars": columns to keep as-is (row identifiers that should NOT be melted).
- Identify "value_vars": columns whose values should be stacked into rows. Leave [] to melt all non-id_vars.
- Identify "var_name": name for the new column holding the former column names.
- Identify "value_name": name for the new column holding the values.
- You may reference intermediate results from the operation history.

Return ONLY the JSON unpivot specification enclosed in $ signs:
${"id_vars": ["col1", "col2"], "value_vars": ["col3", "col4"], "var_name": "key_col", "value_name": "val_col"}$\
"""

AGENT_CODE_GEN = """\
[Code_Gen_And_Score_Tool — get_operator_code_generation_prompt()]

You are generating executable Python code at runtime. Please generate a Python script to convert multiple source tables to the format of the target table and STRICTLY follow the sequence of the operations mentioned in the action history below. The code should be immediately executable in a correct way, which means it should NOT contain any placeholder for brevity. For example, even if there exist hundreds of source tables, these data need to be loaded completely one by one or in a programmable way.

    Transformation Context:

    {query}
    (includes: Target Table Name, Target Schema, Target Examples, Source Information with file locations, CSV save path)

    Action History (follow this sequence of operations strictly):
    {memory_actions}

    Based on the transformation context and action history, generate the Python script that implements the transformation. The script should handle data import, transformation, and export. The script should be complete and executable, not omitting any single statement.

    Please quote the Python script between one single "```Python" and "```".

    Hints to be considered for Python code generation:
[Hints from PYTHON_SCRIPT_HINT_IDS]

 Please quote the Python script between one single "```Python" and "```".

 Errors in previous attempts: {error_string}\
"""

# ── Agentflow orchestration prompts (Planner / Verifier) ───────────────────

AGENT_ANALYZE_QUERY_OPERATOR = """\
[PLANNER — models/planner.py → _analyze_query_operator()]
Called ONCE at the start of each solve attempt. Uses llm_engine_fixed (smaller model).
Returns QueryAnalysis — guides subsequent generate_next_step_operator() calls.

Task: Analyze the given query to determine necessary skills and tools.

CRITICAL CONSTRAINT: You can ONLY suggest operations and tools that are in the available tools list. Do NOT suggest operations or tools that are not available.

Inputs:
- Query: {question}
- Available tools: {self.available_tools}
- Metadata for tools: {self.toolbox_metadata}

Instructions:
1. Identify the main objectives in the query.
2. List ONLY the skills and tools that are actually available in the available tools list.
3. For each skill and tool, explain how it helps address the query.
4. If the query requires operations not supported by available tools, explicitly note this limitation.
5. Note any additional considerations.

Format your response with a summary of the query, lists of skills and tools with explanations, and a section for additional considerations.

IMPORTANT: Be brief and precise. Focus ONLY on what can be achieved with the available tools.\
"""

AGENT_VERIFIER_OPERATOR = """\
[VERIFIER — models/verifier.py → verificate_context() — non-multimodal branch]
Called after every planner → tool cycle. Uses llm_engine_fixed (smaller model).
Returns MemoryVerification(analysis, stop_signal) → STOP or CONTINUE.

Task: Evaluate whether the operator-driven transformation pipeline has produced a correct and complete result, or whether more steps are needed.

Context:
- Query: {question}
- Available Tools: {self.available_tools}
- Toolbox Metadata: {self.toolbox_metadata}
- Initial Analysis: {query_analysis}
- Memory (Tools Used & Results): {memory.get_actions()}

Instructions:
1. Review the query, the operator configurations in memory (Join, Union, GroupBy, Pivot, Unpivot), and any code generation results.
2. Identify whether all necessary operators have been configured to produce the target schema.
3. Code and Score Check — Look for results from Code_Gen_And_Score_Tool in memory:
   - score = 1.0 AND score_summary shows no missed functional dependencies, no missed ground-truth keys, and no missed target columns → conclude STOP.
   - score < 1.0 OR missed items in any category → conclude CONTINUE.
   - No Code_Gen_And_Score_Tool result in memory yet → conclude CONTINUE.
   - execution_success = False in the last result → conclude CONTINUE.

Final Determination:
- Conclude STOP only when Code_Gen_And_Score_Tool has been run AND score = 1.0 with no missed items.
- Conclude CONTINUE in all other cases (operators still missing, code not yet generated, score < 1.0, or execution failed).

IMPORTANT: Keep your response brief (2-3 sentences max). The response must end with either "Conclusion: STOP" or "Conclusion: CONTINUE".\
"""

# ── LangGraph/MCTS prompts ─────────────────────────────────────────────────

LG_EXPAND_FULL = """\
[mcts_expand.py — get_mcts_expand_prompt()]
NOTE: LangGraph COMBINES operator selection + configuration into ONE prompt.

You are generating a data-pipeline to transform multiple source tables to the target table and you need to answer "what operation should be performed next?". Take this decision based on "Operation History", the schema of the source, target tables, and examples in the target table.

Your task: propose up to {k} ALTERNATIVE candidates for the SINGLE NEXT operation step, ranked from most to least promising.

CRITICAL: Each candidate is an INDEPENDENT ALTERNATIVE for the same next step.
They are NOT sequential — candidate 2 does NOT build on candidate 1.
Each candidate must reference ONLY the original source tables (or the last result of
the Operation History), never the output of another candidate.

For EVERY candidate you must specify BOTH the operator type AND its full configuration — no
separate configuration step will be performed.

Allowed Operations: {allowed_operation_list}
Operation History (completed so far): {operation_history}

1. Target Table Name:   {target_data_name}
2. Target Schema:       {target_data_schema}
3. Target Examples:     {target_samples}
4. Source Information:  {source_information}
{fd_hints}

══════════════════════════════════════════════════════
CONFIGURATION RULES — one section per operator type
══════════════════════════════════════════════════════

JOIN — join two tables on shared columns
[Hints from JOIN_HINT_IDS]
  • You should only use columns that actually exist in the source tables.
  Format:
    TABLES: [[table1, table2]]
    COLUMNS: [[table1.col_a, table2.col_b], ...]

UNION — stack tables that have IDENTICAL schemas
  • Only union tables with EXACTLY the same column names; do NOT rename.
  • If tables have different schemas, use JOIN instead of UNION.
  Format:
    TABLES: [table1, table2, ...]

GROUP_BY/AGGREGATE — group rows and apply aggregate functions
[Hints from GROUPBY_AGG_HINT_IDS]
  Format:
    GROUP_BY: [table.col1, ...]
    AGGREGATIONS: [COUNT(table.col), SUM(table.col2), ...]

PIVOT / UNPIVOT — no additional configuration needed
  Format: (just the operator line; no TABLES or COLUMNS line)

NO_MORE_OPERATION — the pipeline is complete
  Format: (just the operator line)

══════════════════════════════════════════════════════
SELECTION GUIDANCE (apply these rules when ranking candidates)
══════════════════════════════════════════════════════
[Hints from NEXT_OPERATOR_HINT_IDS]
- If the operation history already covers all needed transformations → propose NO_MORE_OPERATION.
- Do NOT repeat an operation+configuration already present in the operation history.

══════════════════════════════════════════════════════
OUTPUT FORMAT  (follow exactly)
══════════════════════════════════════════════════════
List up to {k} candidates ranked most-to-least promising.

$CANDIDATE 1$
OPERATOR: <OPERATOR_TYPE>
<configuration lines>
$END$

$CANDIDATE 2$
...
$END$

Now provide your ranked candidates:\
"""

LG_OPERATOR_SELECTION = """\
[Part of mcts_expand.py — SELECTION GUIDANCE section]
NOTE: In LangGraph/MCTS, operator selection and configuration are a single combined prompt.
The selection guidance is embedded within the expansion prompt.

You are generating a data-pipeline to transform multiple source tables to the target table and you need to answer "what operation should be performed next?". Take this decision based on "Operation History", the schema of the source, target tables, and examples in the target table.

SELECTION GUIDANCE (apply these rules when ranking candidates):
[Hints from NEXT_OPERATOR_HINT_IDS]
- If the operation history already covers all needed transformations → propose NO_MORE_OPERATION.
- Do NOT repeat an operation+configuration already present in the operation history.

Context provided:
- Allowed Operations: {allowed_operation_list}
- Operation History: {operation_history}
- Target Table Name, Schema, Examples
- Source Information
- {fd_hints} (functional-dependency hints)

Output: Up to {k} ranked candidates, each with BOTH operator type AND full configuration.\
"""

LG_JOIN = """\
[Part of mcts_expand.py — JOIN configuration section]
NOTE: LangGraph handles JOIN config within the combined expand prompt. No separate JOIN call.

JOIN — join two tables on shared columns
[Hints from JOIN_HINT_IDS]
  • You should only use columns that actually exist in the source tables.
  Format:
    TABLES: [[table1, table2]]
    COLUMNS: [[table1.col_a, table2.col_b], [table1.col_c, table2.col_d], ...]
    (one pair per join condition; use multiple pairs for composite keys)\
"""

LG_UNION = """\
[Part of mcts_expand.py — UNION configuration section]
NOTE: LangGraph handles UNION config within the combined expand prompt. No separate UNION call.

UNION — stack tables that have IDENTICAL schemas
  • Only union tables with EXACTLY the same column names; do NOT rename.
  • If tables have different schemas, use JOIN instead of UNION.
  Format:
    TABLES: [table1, table2, table3, ...]\
"""

LG_GROUPBY = """\
[Part of mcts_expand.py — GROUP_BY/AGGREGATE configuration section]
NOTE: LangGraph handles GROUP_BY config within the combined expand prompt. No separate call.

GROUP_BY/AGGREGATE — group rows and apply aggregate functions
[Hints from GROUPBY_AGG_HINT_IDS]
  Format:
    GROUP_BY: [table.col1, table.col2, ...]
    AGGREGATIONS: [COUNT(table.col), SUM(table.col2), AVG(table.col3), ...]\
"""

LG_PIVOT = """\
[Part of mcts_expand.py — PIVOT/UNPIVOT section]
NOTE: LangGraph includes PIVOT in the combined expand prompt with minimal config required.

PIVOT / UNPIVOT — no additional configuration needed
  Format: (just the operator line; no TABLES or COLUMNS line)

The LLM selects PIVOT as a candidate operator type. Actual PIVOT parameters
(index, columns, values, aggfunc) are resolved at code generation time in mcts_simulate.py.\
"""

LG_UNPIVOT = """\
[Part of mcts_expand.py — PIVOT/UNPIVOT section]
NOTE: LangGraph includes UNPIVOT in the combined expand prompt with minimal config required.

PIVOT / UNPIVOT — no additional configuration needed
  Format: (just the operator line; no TABLES or COLUMNS line)

The LLM selects UNPIVOT as a candidate operator type. Actual UNPIVOT parameters
(id_vars, value_vars, var_name, value_name) are resolved at code generation time in mcts_simulate.py.\
"""

LG_CODE_GEN = """\
[mcts_simulate.py — get_mcts_simulate_prompt()]

You are generating executable Python code at runtime. Please generate a Python script to convert multiple source tables to the format of the target table. The code should be immediately executable in a correct way, which means it should NOT contain any placeholder for brevity. For example, even if there exist hundreds of source tables, these data need to be loaded completely one by one or in a programmable way.

Partial Operation Plan (treat this as a STARTING GUIDE, not a fixed sequence):
{operation_history}

Use the operations above as a hint for the first transformation step(s), then reason
about what additional steps are required to produce the correct target schema.

1. Target Table Name:              {target_data_name}
2. Target Schema:                  {target_data_schema}
3. Target Examples:                {target_samples}
4. Source Information (with paths): {source_information_with_location}
5. Save the result to:             {csv_save_path}

Before writing code, THINK STEP BY STEP about the complete transformation plan:
  a) What does the partial plan above imply for the first operation(s)?
  b) After those operations, what intermediate schema is produced?
  c) What further operations are needed to reach the target schema?
  d) Are there any data-type fixes, string conversions, or column renames required?

Then output your COMPLETE operation plan using this block (one operation per line,
ending with NO_MORE_OPERATION):

$PLAN$
<operation 1>
...
NO_MORE_OPERATION
$END_PLAN$

Then generate the Python script that implements the COMPLETE transformation.

The code must be immediately executable — no placeholders, no ellipses.
Keep the code brief: no inline comments, no docstrings, no explanatory print statements.
Please quote the Python script between one single "```Python" and "```".

Hints for Python code generation:
[Hints from PYTHON_SCRIPT_HINT_IDS]

Errors in previous attempts: {error_string}\
"""

# ── Critique prompts ───────────────────────────────────────────────────────

MULTI_CRITIQUE_HISTORY = """\
[prompts/history_critique.txt — loaded by methods/critique.py]

You are generating a data-pipeline to transform multiple source tables to follw the schema of the target table. Up to this point, an LLM agent has created a Python Code based on Source Information, Target Information. The python script generated didn't work properly. You will be given Target Data Examples, Functional Dependencies discovered in the Target Table, the operators used in the failed Python code, the schema and the examples of the data generated using the failed Python code.

Your task is to select operators and configure each operator, and generate Python code based on the list of selected operators.

--For Join, give the names of tables to be joined and the join attributes;
--For Union, give the names of tables to be unioned;
--For GroupBy, give the Group By attributes; You should try group the data by a couple of LEFTMOST column(s) of the given target table schema. These columns often satisfy the following properties: (1) At the leftmost part of the given target table schema; (2) Must be string, or integer types in the source tables or the given target examples and must NOT be float or decimal types; (3) The column must contain unique (non-redundant) values in the given target table examples.
--For Aggregate, give the aggregation function and aggregation column. Aggregation columns must not be included in group by columns. If many columns in the target table have similar integer values, it probably suggests a count aggregation should be used. If a column (such as X) has float values (such as 1211.2234 or 33.17) in the given target examples, while having integer values (such as 1001 or 35) in the given source tables, it suggests that an "average" aggregation should be applied to X, no matter how X is named. Importantly, this column X MUST BE EXCLUDED from the group by columns.
--For Filter, give the predicate;
--For String operators, give a description;
--It may also contain other operators, such as Complete Table (Add missing rows to the table), Pivot, Next Row (Give the next row).

Source Data Information: $SRC_INFO$
Target Data Information :
Schema :  $SCHEMA$
Number of Tuples: $NUM_TUPLES$
Examples : $EXAMPLES$

Functional Dependencies (FDs) (Important FDs are ranked on top):
$FD_HINT$

The failed Python script follow the following operations: $OPERATIONS$
After these operations, the resulting table that doesn't match target data has schema and examples as follows:
Schema : $RES_SCHEMA_NOT_AVAILABLE$
Number of Tuples: $NUM_RES_TUPLES_NOT_AVAILABLE$
Examples : $RES_EXAMPLES_NOT_AVAILABLE$

Now you can change those operations to fix the failed transformation. Your code should only take the CSV file paths given in the Source Data Information as inputs. Please make use of all source tables. Do not miss any source table. Please pay attention to the following hints and select relevant hints to apply.

$STATIC_HINTS$
[Hints from CRITIQUE_HINT_IDS — injected when static_hints flag is on]

$METADATA$

$FEW_SHOT_EXAMPLES$\
"""

MULTI_CRITIQUE_CODE_REFINE = """\
[Second LLM call in methods/critique.py — query_generator]
NOTE: This is a follow-up call that takes the critique response and refines the Python code.

Based on the LLM response, can you refine the python code.

[Optional hints from CRITIQUE_HINT_IDS]

Note : - Make sure to write the final output of the python code to {target_location_critique}
    - Make sure to write the python code in-between "```Python" and "```"
    - Please keep the final output columns the same as it was in the python script given. [Strictly do not add prefix or suffix to the column names]
    - You just need to apply the fix according to the criticizer response.
    - Do not use assignment operation for any column.
    Python Code : ```Python
    {python_code}
    ```

    Code fixing suggestions : ```
    {res}
    ```\
"""

AGENT_CRITIQUE = """\
[Critique_Pipeline_Tool — get_critique_prompt()]

You are critiquing a data-pipeline that transforms multiple source tables to follow the schema of the target table. An LLM agent has been building this pipeline step by step. Review the current pipeline and suggest corrections.

Your task is to review the current operation history and suggest what operators need to be changed, added, or removed to correctly produce the target table.

{transformation_case_info}
(includes: Target Table Name, Target Schema, Target Examples, Source Information, Operation History)

For each operator you suggest, provide clear configuration:
--For Join, give the names of tables to be joined and the join attributes.
--For Union, give the names of tables to be unioned.
--For GroupBy, give the Group By attributes. Group by a couple of LEFTMOST column(s) of the target table schema. These columns must be string or integer types (NOT float) and contain unique values in the target examples.
--For Aggregate, give the aggregation function and aggregation column. Aggregation columns must not be included in group by columns.

Please pay attention to the following hints and apply relevant ones:

[Hints from CRITIQUE_HINT_IDS — numbered]

Based on your analysis, provide:
1. What is wrong with the current pipeline (if anything).
2. What specific operators and configurations should be added, changed, or removed.
3. The corrected sequence of operations.

Format your response as:
$CRITIQUE: <your detailed critique and corrected pipeline>$\
"""

LG_CRITIQUE = "(Not applicable — LangGraph/MCTS does not have a dedicated critique step. The MCTS tree search explores alternative operator candidates instead of critiquing a fixed pipeline.)"

# ── Pipeline-level prompt texts ────────────────────────────────────────────

MULTI_SINGLE_STEP_COT = """\
[prompts/code_generation_prompt.py → get_python_script_for_single_step_cot()]

You are generating executable Python code at runtime. Please generate a Python script to convert multiple source tables to the format of the target table. The code should immediately executable in a correct way, which means it should NOT contain any placeholder for brievity. For example, even if there exists hundreds of source tables, these data needs to be loaded completely one by one or in a programmable way. Before generating the code, please think step by step about the transformation plan to convert the source tables to the target table.

    1. Target Table Name: {target_data_name}
    2. Target Schema: {target_data_schema}
    3. Target Examples: {target_samples}
    4. Source Information: {source_information_with_location}
    5. Write the result to this path {csv_save_path}

    Based on the transformation plan, generate the Python script that implements the transformation. The script should handle data import, transformation, and export. The script should be complete and executable, not omiting any single statement. For example, please list all the source paths that will be used.

    Please quote the Python script between one single "```Python" and "```".

Hints to be considered for Python code generation:
[Hints from PIPELINE_HINT_IDS — numbered, always included]

Please quote the Python script between one single "```Python" and "```".

Errors in previous Attempts : {error_string}\
"""

AGENT_CREATE_PIPELINE = """\
[AgentFlow/.../tools/Create_New_Pipeline/tool.py → get_pipeline_prompt()]
NOTE: Single LLM call — pipeline design + code generation combined.

You are generating executable Python code at runtime. Please generate a Python script to convert multiple source tables to the format of the target table. The code must be immediately executable in a correct way, which means it should NOT contain any placeholder for brevity. For example, even if there exist hundreds of source tables, these data need to be loaded completely one by one or in a programmable way. Before generating the code, think step by step about the full transformation plan.

{transformation_case_info}
(includes: Target Table Name, Target Schema, Target Examples, Source Information)

STEP 1 — Think step by step about the full pipeline:
  a) What operations (JOIN/UNION/GROUP_BY/AGGREGATE/PIVOT/UNPIVOT) are needed and in what order?
  b) What columns appear in the target but need aggregation from source?
  c) Are all source tables used?

STEP 2 — Output the complete pipeline plan:
$PIPELINE:
Step 1: <OPERATION_TYPE> - <configuration details>
Step 2: <OPERATION_TYPE> - <configuration details>
...
$

STEP 3 — Generate the executable Python script implementing the plan above.
The code must be immediately executable — no placeholders. Load all source CSV files listed in the source information.
Please quote the Python script between one single "```Python" and "```".

[Hints from PIPELINE_HINT_IDS — numbered]\
"""

AGENT_REFINE_PIPELINE = """\
[AgentFlow/.../tools/Refine_Existing_Pipeline/tool.py → get_refine_pipeline_prompt()]
NOTE: Single LLM call — self-contained critique + corrected plan + corrected code.

You are critiquing a data-pipeline that transforms multiple source tables to follow the schema of the target table. An LLM agent has created this pipeline plan and generated Python code. Review the current pipeline and suggest corrections, then provide the corrected pipeline plan and executable code.

{transformation_case_info}
(includes: Target Table Name, Target Schema, Target Examples, Source Information)

Current Pipeline:
{current_pipeline}

[Execution Context if provided — score feedback, error messages, intermediate output]

Your task:
1. Identify what is wrong with the current pipeline (if anything).
2. Specify what steps need to be added, changed, or removed.
3. Output the CORRECTED pipeline plan.
4. Generate the executable Python script for the corrected pipeline. The code must be immediately executable — no placeholders.

[Hints from CRITIQUE_HINT_IDS — numbered]

Output format:
$CRITIQUE: <your detailed analysis of what is wrong and what should change>$

$PIPELINE:
Step 1: <OPERATION_TYPE> - <configuration details>
Step 2: <OPERATION_TYPE> - <configuration details>
...
$

Please quote the Python script between one single "```Python" and "```".\
"""

AGENT_PLANNER_ANALYZE_QUERY_PIPELINE = """\
[PLANNER — models/planner.py → _analyze_query_pipeline()]
Called ONCE at the start. Uses llm_engine_fixed (smaller model).
Returns QueryAnalysis — decides CREATE NEW vs MODIFY EXISTING pipeline.

Task: Analyze the given query at a PIPELINE level to determine whether a new data transformation pipeline needs to be created or an existing pipeline should be modified.

CRITICAL CONSTRAINT: You can ONLY suggest operations and tools that are in the available tools list. Do NOT suggest operations or tools that are not available.

Inputs:
- Query: {question}
- Available tools: {self.available_tools}
- Metadata for tools: {self.toolbox_metadata}

Instructions:
1. Understand the high-level goal of the query — what is the desired end-to-end data transformation?
2. Determine whether this requires:
   a) Creating a NEW pipeline from scratch (no existing pipeline addresses this transformation), or
   b) Modifying an EXISTING pipeline (the transformation is partially addressed and needs adjustments).
3. Choose the operation sequence to build the target table: which joins, unions, aggregations, etc. steps to apply, in what order.
4. List ONLY the skills and tools that are actually available in the available tools list.
5. Note any additional considerations at the pipeline level.

Format your response with:
- A concise summary of the query and the target transformation
- Pipeline decision: CREATE NEW or MODIFY EXISTING (and why)
- Required skills and relevant tools from the available list
- Additional considerations

IMPORTANT: Be brief and precise. Focus on the pipeline-level strategy, NOT individual operator details.\
"""

AGENT_PLANNER_NEXT_STEP_PIPELINE = """\
[PLANNER — models/planner.py → _generate_next_step_pipeline()]
Called at each loop iteration. Selects Create_New_Pipeline or Refine_Existing_Pipeline.
Returns NextStep(justification, context, sub_goal, tool_name).

Task: Determine the optimal next step at a PIPELINE level to address the given query.

You are operating at a PIPELINE granularity. Instead of picking individual operators one at a time, you should reason about the pipeline as a whole and decide whether to:
1. Create a new pipeline — Design the full sequence of operations needed to achieve the target transformation.
2. Modify the existing pipeline — Adjust, add, remove, or reconfigure operations to fix issues or improve correctness.

CRITICAL CONSTRAINT: You can ONLY use tools from the available tools list. Do NOT suggest operations, sub-goals, or actions that require tools not in this list.

Context:
- Query: {question}
- Query Analysis: {query_analysis}
- Available Tools (THESE ARE THE ONLY TOOLS YOU CAN USE): {self.available_tools}
- Toolbox Metadata: {self.toolbox_metadata}
- Current Step: {step_count} of {max_step_count}
- Remaining Steps: {max_step_count - step_count}
- Previous Steps and Their Results: {memory.get_actions()}

Instructions:
1. Review the query, the pipeline-level analysis, and all previous steps.
2. Decide the pipeline-level action:
   - If a pipeline is in progress, evaluate whether it is on track or needs refinement.
3. Select the single best tool from the available tools list for the next step.
4. Formulate a specific, achievable sub-goal for the selected tool.
5. Provide all necessary context (data, file names, variables) for the tool to function.

Pipeline-Level Reasoning:
- Plan the end-to-end transformation from sources to target.
- When creating a new pipeline, plan the full sequence.
- When modifying, identify exactly what is wrong and what change will fix it.

Response Format:
1. Justification: Explain your pipeline-level reasoning and choice of tool.
2. Context: Provide all necessary information for the tool.
3. Sub-Goal: State the specific objective for the tool.
4. Tool Name: State the exact name of the selected tool (must be from available tools list).

Rules:
- Select only ONE tool from the available tools list.
- The sub-goal must be directly achievable by the selected tool.
- The response must end with the Context, Sub-Goal, and Tool Name sections in that order, with no extra content.\
"""

AGENT_VERIFIER_PIPELINE = """\
[VERIFIER — models/verifier.py → verificate_context_pipeline()]
Called after each pipeline tool. Uses llm_engine_fixed (smaller model).
Returns PipelineVerification(analysis, stop_signal, finalized_pipeline_id) → STOP or CONTINUE.

Task: Evaluate if the current pipeline is complete and correct enough to produce the target transformation, or if more modifications are needed.

Context:
- Query: {question}
- Available Tools: {self.available_tools}
- Toolbox Metadata: {self.toolbox_metadata}
- Initial Analysis: {query_analysis}
- Memory (Tools Used & Results): {memory.get_actions()}

Instructions:
1. Review the query and the pipeline(s) created/modified in memory.
2. Assess completeness: Does the latest pipeline fully address the target transformation?
3. Check the critique: If the latest pipeline's critique says CORRECT, the pipeline is likely ready. However, the critique alone is not sufficient — the score must also confirm correctness.
4. Score Understanding — interpret score_summary from memory:
   - Functional Dependency Matching: If functional_dependencies.missed is non-empty → add GROUP BY.
   - Key Matching: If keys.missed_ground_truth_keys is non-empty → add GROUP BY on those columns.
   - Column Mapping: If column_mappings.missed_target_columns is non-empty → add JOINs.
   - The pipeline is ONLY correct when ALL three are satisfied: score = 1.0, no missed items.
5. Score thresholds:
   - overall_score = 1.0 AND all missed lists empty → recommend STOP.
   - Any missed item or low score → recommend CONTINUE.
   - Failed execution → recommend CONTINUE.
6. If the pipeline is complete and correct, identify the pipeline_id of the finalized pipeline from the memory.

Final Determination:
- If pipeline is complete and correct: explain why, provide pipeline_id, conclude STOP.
- If modifications still needed: explain what's missing or incorrect, conclude CONTINUE.

IMPORTANT: The response must end with either "Conclusion: STOP" or "Conclusion: CONTINUE".
If STOP, you MUST include the finalized_pipeline_id with the pipeline_id to finalize.\
"""

LG_PIPELINE_NA = "(Not applicable — LangGraph/MCTS operates at operator granularity only via the MCTS tree search; there is no pipeline-level counterpart.)"

PIPELINE_STEPS = [
    {
        "step": "Planner — Analyze Query",
        "description": "One-time: decide CREATE NEW vs MODIFY EXISTING pipeline before the loop begins",
        "multi_file":  "N/A — single_step_cot makes a direct call without a planning step",
        "multi_prompt": "(Not applicable — Multi-Step single_step_cot does not have a planner/verifier loop; it calls the pipeline design+code prompt once per attempt.)",
        "agent_file":  "AgentFlow/.../models/planner.py → _analyze_query_pipeline()",
        "agent_prompt": AGENT_PLANNER_ANALYZE_QUERY_PIPELINE,
        "lg_file":     "N/A",
        "lg_prompt":   LG_PIPELINE_NA,
    },
    {
        "step": "Planner — Next Step Decision",
        "description": "At each loop iteration: choose Create_New_Pipeline or Refine_Existing_Pipeline",
        "multi_file":  "N/A",
        "multi_prompt": "(Not applicable — Multi-Step single_step_cot does not have a planner loop.)",
        "agent_file":  "AgentFlow/.../models/planner.py → _generate_next_step_pipeline()",
        "agent_prompt": AGENT_PLANNER_NEXT_STEP_PIPELINE,
        "lg_file":     "N/A",
        "lg_prompt":   LG_PIPELINE_NA,
    },
    {
        "step": "Create Pipeline (Design + Code)",
        "description": "Single LLM call — think through full pipeline plan then generate executable code",
        "multi_file":  "prompts/code_generation_prompt.py → get_python_script_for_single_step_cot()",
        "multi_prompt": MULTI_SINGLE_STEP_COT,
        "agent_file":  "AgentFlow/.../tools/Create_New_Pipeline/tool.py → get_pipeline_prompt()",
        "agent_prompt": AGENT_CREATE_PIPELINE,
        "lg_file":     "N/A",
        "lg_prompt":   LG_PIPELINE_NA,
    },
    {
        "step": "Critique + Refine Pipeline",
        "description": "Single LLM call — self-analyze current pipeline, produce corrected plan + code",
        "multi_file":  "N/A — no pipeline-level critique/refine in single_step_cot",
        "multi_prompt": "(Not applicable — Multi-Step single_step_cot does not have a pipeline-level critique/refine step. Critique is at operator level via methods/critique.py.)",
        "agent_file":  "AgentFlow/.../tools/Refine_Existing_Pipeline/tool.py → get_refine_pipeline_prompt()",
        "agent_prompt": AGENT_REFINE_PIPELINE,
        "lg_file":     "N/A",
        "lg_prompt":   LG_PIPELINE_NA,
    },
    {
        "step": "Verifier — STOP / CONTINUE",
        "description": "After each pipeline tool: decide whether the pipeline is correct (STOP) or needs refinement (CONTINUE)",
        "multi_file":  "N/A — no verifier; single_step_cot runs for a fixed number of attempts",
        "multi_prompt": "(Not applicable — Multi-Step single_step_cot does not have a verifier; it reruns the full pipeline prompt on each failed attempt.)",
        "agent_file":  "AgentFlow/.../models/verifier.py → verificate_context_pipeline()",
        "agent_prompt": AGENT_VERIFIER_PIPELINE,
        "lg_file":     "N/A",
        "lg_prompt":   LG_PIPELINE_NA,
    },
]

# ── build sheet data ───────────────────────────────────────────────────────

STEPS = [
    {
        "step": "Planner — Analyze Query",
        "description": "One-time: identify required operators and tools from the query before the loop begins",
        "multi_file":  "N/A — no explicit query-analysis step",
        "multi_prompt": "(Not applicable — Multi-Step does not have a separate query-analysis/planner step; it directly calls each prompt in sequence.)",
        "agent_file":  "AgentFlow/.../models/planner.py → _analyze_query_operator()",
        "agent_prompt": AGENT_ANALYZE_QUERY_OPERATOR,
        "lg_file":     "N/A",
        "lg_prompt":   "(Not applicable — LangGraph/MCTS uses MCTS tree search; there is no separate query-analysis/planner step.)",
    },
    {
        "step": "Operator Selection",
        "description": "Decide which operator to apply next",
        "multi_file":  "prompts/next_operator_prompt.py → get_next_operator_prompt()",
        "multi_prompt": MULTI_OPERATOR_SELECTION,
        "agent_file":  "AgentFlow/.../models/planner.py → _generate_next_step_operator()",
        "agent_prompt": AGENT_OPERATOR_SELECTION,
        "lg_file":     "prompts/mcts_expand.py → get_mcts_expand_prompt() [SELECTION GUIDANCE section]",
        "lg_prompt":   LG_OPERATOR_SELECTION,
    },
    {
        "step": "JOIN Configuration",
        "description": "Determine which tables and columns to join",
        "multi_file":  "prompts/configuration_prompts.py → get_join_prompt()",
        "multi_prompt": MULTI_JOIN,
        "agent_file":  "AgentFlow/.../tools/configure_join_operator/tool.py → get_join_prompt()",
        "agent_prompt": AGENT_JOIN,
        "lg_file":     "prompts/mcts_expand.py → get_mcts_expand_prompt() [JOIN section — combined with selection]",
        "lg_prompt":   LG_JOIN,
    },
    {
        "step": "UNION Configuration",
        "description": "Determine which tables to stack (same schema)",
        "multi_file":  "prompts/configuration_prompts.py → get_union_prompt()",
        "multi_prompt": MULTI_UNION,
        "agent_file":  "AgentFlow/.../tools/configure_union_operator/tool.py → get_union_prompt()",
        "agent_prompt": AGENT_UNION,
        "lg_file":     "prompts/mcts_expand.py → get_mcts_expand_prompt() [UNION section — combined with selection]",
        "lg_prompt":   LG_UNION,
    },
    {
        "step": "GROUP_BY / AGGREGATE Configuration",
        "description": "Determine group-by columns, aggregation columns and functions",
        "multi_file":  "prompts/configuration_prompts.py → get_group_by_aggregate_prompt()",
        "multi_prompt": MULTI_GROUPBY,
        "agent_file":  "AgentFlow/.../tools/configure_group_by_aggregate_operator/tool.py → get_group_by_aggregate_prompt()",
        "agent_prompt": AGENT_GROUPBY,
        "lg_file":     "prompts/mcts_expand.py → get_mcts_expand_prompt() [GROUP_BY section — combined with selection]",
        "lg_prompt":   LG_GROUPBY,
    },
    {
        "step": "PIVOT Configuration",
        "description": "Turn row values into column headers (wide format)",
        "multi_file":  "N/A — no dedicated PIVOT prompt",
        "multi_prompt": MULTI_PIVOT,
        "agent_file":  "AgentFlow/.../tools/add_pivot/tool.py → get_pivot_prompt()",
        "agent_prompt": AGENT_PIVOT,
        "lg_file":     "prompts/mcts_expand.py → PIVOT section (minimal; params resolved at simulation)",
        "lg_prompt":   LG_PIVOT,
    },
    {
        "step": "UNPIVOT Configuration",
        "description": "Turn column headers into rows (narrow format / melt)",
        "multi_file":  "N/A — no dedicated UNPIVOT prompt",
        "multi_prompt": MULTI_UNPIVOT,
        "agent_file":  "AgentFlow/.../tools/add_unpivot/tool.py → get_unpivot_prompt()",
        "agent_prompt": AGENT_UNPIVOT,
        "lg_file":     "prompts/mcts_expand.py → UNPIVOT section (minimal; params resolved at simulation)",
        "lg_prompt":   LG_UNPIVOT,
    },
    {
        "step": "Code Generation",
        "description": "Generate executable Python script for the transformation",
        "multi_file":  "prompts/code_generation_prompt.py → get_python_script_simple()",
        "multi_prompt": MULTI_CODE_GEN,
        "agent_file":  "AgentFlow/.../tools/code_gen_and_score/tool.py → get_operator_code_generation_prompt()",
        "agent_prompt": AGENT_CODE_GEN,
        "lg_file":     "prompts/mcts_simulate.py → get_mcts_simulate_prompt()",
        "lg_prompt":   LG_CODE_GEN,
    },
    {
        "step": "Critique — Pipeline Analysis",
        "description": "Identify what is wrong with the current pipeline and suggest corrections",
        "multi_file":  "prompts/history_critique.txt (loaded by methods/critique.py)",
        "multi_prompt": MULTI_CRITIQUE_HISTORY,
        "agent_file":  "AgentFlow/.../tools/critique_pipeline/tool.py → get_critique_prompt()",
        "agent_prompt": AGENT_CRITIQUE,
        "lg_file":     "N/A",
        "lg_prompt":   LG_CRITIQUE,
    },
    {
        "step": "Critique — Code Refinement",
        "description": "Refine the Python script based on the critique response (Multi-Step only)",
        "multi_file":  "methods/critique.py → query_generator (second LLM call)",
        "multi_prompt": MULTI_CRITIQUE_CODE_REFINE,
        "agent_file":  "N/A — Agentflow critique is a single call; code re-generation is handled by Code_Gen_And_Score_Tool on the next planner step",
        "agent_prompt": "(No separate code-refinement prompt — the planner invokes Code_Gen_And_Score_Tool again after critique feedback is stored in memory)",
        "lg_file":     "N/A",
        "lg_prompt":   LG_CRITIQUE,
    },
    {
        "step": "Verifier — STOP / CONTINUE",
        "description": "After each tool call: decide whether the pipeline is complete (STOP) or needs more operators/critique (CONTINUE)",
        "multi_file":  "N/A — no verifier; Multi-Step runs a fixed-length sequential loop",
        "multi_prompt": "(Not applicable — Multi-Step does not have a verifier. It runs a fixed number of operator-selection → configuration → code-generation cycles.)",
        "agent_file":  "AgentFlow/.../models/verifier.py → verificate_context() (non-multimodal)",
        "agent_prompt": AGENT_VERIFIER_OPERATOR,
        "lg_file":     "N/A",
        "lg_prompt":   "(Not applicable — LangGraph/MCTS determines termination via simulation scores in the MCTS tree; there is no dedicated verifier prompt.)",
    },
]

# ── helpers ────────────────────────────────────────────────────────────────

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def make_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

def write_cell(ws, row, col, value, fill=None, font=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:      cell.fill      = fill
    if font:      cell.font      = font
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    return cell

# ── main builder ──────────────────────────────────────────────────────────

def build_sheet(wb):
    # Remove existing sheet if present
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(title=SHEET_NAME)

    # Column layout:
    # A: Step / label (narrow)        col 1
    # B: Multi-Step prompt             col 2
    # C: Agentflow prompt              col 3
    # D: LangGraph/MCTS prompt         col 4

    # Set column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 70

    hdr_fill  = make_fill(C_HEADER_BG)
    hdr_font  = make_font(bold=True, color=C_HEADER_FG, size=11)
    sect_fill = make_fill(C_SECTION_BG)
    sect_font = make_font(bold=True, size=11)
    lbl_font  = make_font(bold=True, size=9)
    body_font = make_font(size=9)
    border    = make_border()

    row = 1

    # ── title row ─────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 30
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value="Prompt Design Comparison — Operator Granularity")
    c.font      = make_font(bold=True, color=C_HEADER_FG, size=14)
    c.fill      = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 1

    # ── sub-title / legend ────────────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1,
                value=("Each section shows one operator/step. "
                       "Columns B-D show the prompt text used by each design for that step. "
                       "{placeholders} are runtime variables filled from source/target info."))
    c.font      = make_font(size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    row += 1

    row += 1  # blank row

    # ── column header row ─────────────────────────────────────────────────
    ws.row_dimensions[row].height = 36
    headers = ["Step / Label", "Multi-Step (Sequential CoT)", "Agentflow (Operator Granularity)", "LangGraph / MCTS"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
    row += 1

    # ── freeze panes ──────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=row, column=1)

    # ── per-step sections ─────────────────────────────────────────────────
    for step_data in STEPS:
        step_name   = step_data["step"]
        description = step_data["description"]

        # Section banner row
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=f"▶  {step_name}   —   {description}")
        c.font      = sect_font
        c.fill      = sect_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = border
        row += 1

        # Source file row
        ws.row_dimensions[row].height = 30
        labels   = ["Source file / function"]
        files    = [step_data["multi_file"], step_data["agent_file"], step_data["lg_file"]]
        fills    = [make_fill(C_MULTI_BG), make_fill(C_AGENT_BG), make_fill(C_LG_BG)]
        for col_idx, (val, fill) in enumerate(
                [("Source file / function", make_fill(C_NOTE_BG))] +
                list(zip(files, fills)),
                start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font      = lbl_font
            c.fill      = fill
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border    = border
        row += 1

        # Prompt text row  (tall)
        prompts = [step_data["multi_prompt"], step_data["agent_prompt"], step_data["lg_prompt"]]
        # estimate height: roughly 15pt per line, but capped at 400
        max_lines = max(p.count("\n") for p in prompts) + 2
        row_height = min(max(max_lines * 13, 80), 600)
        ws.row_dimensions[row].height = row_height

        write_cell(ws, row, 1, "Prompt text",
                   fill=make_fill(C_NOTE_BG), font=lbl_font,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                   border=border)
        for col_idx, (prompt, fill) in enumerate(zip(prompts, fills), start=2):
            write_cell(ws, row, col_idx, prompt,
                       fill=fill, font=body_font,
                       alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                       border=border)
        row += 1

        # blank separator
        row += 1

    # ── auto-filter on header row ─────────────────────────────────────────
    # (not very meaningful here, but handy)
    # ws.auto_filter.ref = f"A5:D5"

    print(f"Sheet '{SHEET_NAME}' built with {row} rows.")
    return ws


def build_pipeline_sheet(wb):
    # Remove existing sheet if present
    if PIPELINE_SHEET_NAME in wb.sheetnames:
        del wb[PIPELINE_SHEET_NAME]

    ws = wb.create_sheet(title=PIPELINE_SHEET_NAME)

    # Column layout:
    # A: Step / label                               col 1
    # B: Multi-Step (single_step_cot)               col 2
    # C: Agentflow — Create_New_Pipeline            col 3
    # D: Agentflow — Refine_Existing_Pipeline       col 4

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 70

    hdr_fill  = make_fill(C_HEADER_BG)
    hdr_font  = make_font(bold=True, color=C_HEADER_FG, size=11)
    sect_fill = make_fill(C_SECTION_BG)
    sect_font = make_font(bold=True, size=11)
    lbl_font  = make_font(bold=True, size=9)
    body_font = make_font(size=9)
    border    = make_border()

    row = 1

    # ── title row ─────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 30
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value="Prompt Design Comparison — Pipeline Granularity")
    c.font      = make_font(bold=True, color=C_HEADER_FG, size=14)
    c.fill      = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 1

    # ── sub-title / legend ────────────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1,
                value=("Each section shows one pipeline-level step. "
                       "Multi-Step uses single_step_cot (one combined call). "
                       "Agentflow has Create_New_Pipeline (design+code) and Refine_Existing_Pipeline (critique+correct). "
                       "LangGraph/MCTS has no pipeline-level equivalent."))
    c.font      = make_font(size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    row += 1

    row += 1  # blank row

    # ── column header row ─────────────────────────────────────────────────
    ws.row_dimensions[row].height = 36
    headers = [
        "Step / Label",
        "Multi-Step (single_step_cot)",
        "Agentflow — Create_New_Pipeline",
        "Agentflow — Refine_Existing_Pipeline",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
    row += 1

    # ── freeze panes ──────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=row, column=1)

    # Column D uses a distinct colour (teal) to distinguish Refine from Create
    C_REFINE_BG = "D9EAD3"  # slightly different green

    # ── per-step sections ─────────────────────────────────────────────────
    for step_data in PIPELINE_STEPS:
        step_name   = step_data["step"]
        description = step_data["description"]

        # Section banner row
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=f"▶  {step_name}   —   {description}")
        c.font      = sect_font
        c.fill      = sect_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = border
        row += 1

        # Source file row
        ws.row_dimensions[row].height = 30
        files = [step_data["multi_file"], step_data["agent_file"], step_data["lg_file"]]
        fills = [make_fill(C_MULTI_BG), make_fill(C_AGENT_BG), make_fill(C_REFINE_BG)]
        for col_idx, (val, fill) in enumerate(
                [("Source file / function", make_fill(C_NOTE_BG))] +
                list(zip(files, fills)),
                start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font      = lbl_font
            c.fill      = fill
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border    = border
        row += 1

        # Prompt text row (tall)
        prompts = [step_data["multi_prompt"], step_data["agent_prompt"], step_data["lg_prompt"]]
        max_lines = max(p.count("\n") for p in prompts) + 2
        row_height = min(max(max_lines * 13, 80), 600)
        ws.row_dimensions[row].height = row_height

        write_cell(ws, row, 1, "Prompt text",
                   fill=make_fill(C_NOTE_BG), font=lbl_font,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                   border=border)
        for col_idx, (prompt, fill) in enumerate(zip(prompts, fills), start=2):
            write_cell(ws, row, col_idx, prompt,
                       fill=fill, font=body_font,
                       alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                       border=border)
        row += 1

        # blank separator
        row += 1

    print(f"Sheet '{PIPELINE_SHEET_NAME}' built with {row} rows.")
    return ws


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    build_sheet(wb)
    build_pipeline_sheet(wb)
    wb.save(XLSX_PATH)
    print(f"Saved → {XLSX_PATH}")


if __name__ == "__main__":
    main()
